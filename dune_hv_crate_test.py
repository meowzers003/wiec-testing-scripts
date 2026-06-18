import pyvisa
import sys
import json
import os
import time
import openpyxl
from datetime import datetime
from keysight_daq970a import Keysight970A
from rigol_dp832a import RigolDP832A
from caen_r8033dm_wrapper import CAENR8033DM_WRAPPER

import csv
from pathlib import Path
import matplotlib.dates as mdates
import matplotlib.ticker as mticker
from datetime import datetime, timedelta
import matplotlib.pyplot as plt
from matplotlib.ticker import MultipleLocator
import numpy as np
from scipy.optimize import curve_fit

import traceback

class LDOmeasure:
    def __init__(self, config_file = None, name = None):
        self.prefix = "DUNE HV Crate Tester"
        print(f"{self.prefix} --> Welcome to the DUNE HV crate production testing script")
        if not config_file:
            print(f"{self.prefix} --> No config file given, test will not run")
            return
        with open(config_file, "r") as jsonfile:
            self.json_data = json.load(jsonfile)
        self.rm = pyvisa.ResourceManager('@py')      
        
        
        #Initialize all instruments first so that you don't waste time with input if something is not connected
        self.c = CAENR8033DM_WRAPPER(self.json_data)
        self.k = Keysight970A(self.rm, self.json_data)

        #Since there are 2 Rigols, set them up here so they know what channels they have
        #And if the test sequence calls the wrong one, it'll throw an error
        self.r0 = RigolDP832A(self.rm, self.json_data, 0)
        self.r0.setup_fan()
        self.r0.setup_heater_supply()
        self.r0.setup_heater_switch()

        self.r1 = RigolDP832A(self.rm, self.json_data, 1)
        self.r1.setup_hvpullup()
        self.r1.setup_hvpullup2()
        self.r1.setup_fanread()
        #Now we can get the input for the name of the test
        if (name):
            self.test_name = name
        else:
            self.test_name = input("Input the test name:\n")

        self.rounding_factor = self.json_data["rounding_factor"]
        #The datastore is the eventual output JSON file that will be written after the test
        #Want to also include what the inputs for this particular test was
        self.datastore = {}
        self.datastore['input_params'] = self.json_data
        self.datastore['test_name'] = self.test_name
        self.start_time = datetime.now()
        self.datastore['start_time'] = self.start_time
        self.initialize_spreadsheet()

        self.fan_test_result = True
        self.heat_test_result = True
        self.hv_test_result = True

        self.datastore['Tests'] = {}
        
        # Run analytical RC filter model before fan, fan PWM, heater, and HV tests.
        self.store_rc_filter_time_constants()

        try:
            self.fan_test()
            self.wb.save(self.path_to_spreadsheet)

            self.fan_test_sweep() # trigger PWN measurement fan test
            self.wb.save(self.path_to_spreadsheet)

            self.heater_test()
            self.wb.save(self.path_to_spreadsheet)

        

            self.hv_test() #add'l specific exceptions are handled within 
            
        except:
            print("Detected exception, powering off all devices first.")
            self.emergency_shutoff()		
            raise

        self.emergency_shutoff() #just in case not already off  	
	
        if (self.fan_test_result and self.heat_test_result and self.hv_test_result):
            self.ws.cell(row=self.row, column=1, value=self.test_name).style = "pass"
            self.datastore['overall'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=1, value=self.test_name).style = "fail"
            self.datastore['overall'] = "Fail"
        self.wb.save(self.path_to_spreadsheet)

        end_time = datetime.now()
        test_time = end_time - self.start_time
        self.datastore['end_time'] = end_time
        self.datastore['test_time'] = test_time

        with open(self.json_output_file, 'w', encoding='utf-8') as f:
            json.dump(self.datastore, f, ensure_ascii=False, indent=4, default=str)
            
 	     
        print(f"{self.prefix} --> Test complete")
        print(f"{self.prefix} --> Test result: {self.datastore['overall']}")
        self.beep_sequence()
        #self.make_hv_plots()

    #Looks to see if a main spreadsheet of all results exists. If it does, open it and find the next row to write these results to
    #If not, it creates the spreadsheet with the proper headers and formatting
    def initialize_spreadsheet(self):
        #In the config JSON I give the user the ability to choose the absolute path to dump all this stuff into, or make it relative to this Python script
        if (self.json_data["relative"] == "True"):
            output_path = os.path.abspath(self.json_data["output_directory"])
        else:
            output_path = os.path.normpath(self.json_data["output_directory"])

        self.path_to_spreadsheet = os.path.join(output_path, f"{self.json_data['output_file']}")
        self.datastore['spreadsheet_path'] = self.path_to_spreadsheet

        json_date = datetime.today().strftime('%Y%m%d%H%M%S')
        #json_date = "20240125152313"
        os.makedirs(os.path.join(output_path, json_date))
        self.results_path = os.path.join(output_path, json_date)
        self.json_output_file = os.path.join(self.results_path, f"{json_date}_{self.test_name}.json")
        self.datastore['json_path'] = self.json_output_file

        self.hv_cols = 12

        # Existing fan_test() result columns:
        # col 4 = Supply Voltage
        # col 5 = Supply Current
        # col 6 onward = Fan RD results
        self.fan_rd_first_col = 6

        # New Fan PWM sweep columns are inserted right after Fan RD results.
        self.fan_pwm_first_col = self.fan_rd_first_col + 6
        self.fan_pwm_num_fans = int(self.json_data['keysight970a_fan_num'])
        self.fan_pwm_cols = 3 + self.fan_pwm_num_fans

        # Shift Heater and HV sections to the right to make room for Fan PWM data.
        self.tc_res_first_col = self.fan_pwm_first_col + self.fan_pwm_cols
        self.hv_res_first_col = self.tc_res_first_col + 8     

        if (os.path.isfile(self.path_to_spreadsheet)):
            self.wb = openpyxl.load_workbook(filename = self.path_to_spreadsheet)
            self.ws = self.wb.active

            # existing output_file may not have Fan PWM columns yet.
            # if missing, insert them between Fan RD results and Heater results.
            self.ensure_fan_pwm_columns()

            self.row = self.ws.max_row + 1

        else:
            self.wb = openpyxl.Workbook()
            self.ws = self.wb.active
            self.ws.title = self.json_data["sheet_name"]

            #Once styles are added, they are saved even if you close and reopen the spreadsheet
            top_style = openpyxl.styles.NamedStyle(name="top")
            top_style.font = openpyxl.styles.Font(bold=True, name='Calibri')
            bd = openpyxl.styles.Side(style='thin')
            top_style.border = openpyxl.styles.Border(left=bd, top=bd, right=bd, bottom=bd)
            self.wb.add_named_style(top_style)

            self.fail_style = openpyxl.styles.NamedStyle(name="fail")
            self.fail_style.font = openpyxl.styles.Font(color="FF0000")
            self.wb.add_named_style(self.fail_style)

            self.pass_style = openpyxl.styles.NamedStyle(name="pass")
            self.pass_style.font = openpyxl.styles.Font(color="009900")
            self.wb.add_named_style(self.pass_style)

            #Yes I'm using magic numbers, but I need to build the spreadsheet in a specific way, hopefully it's only done once here
            ##nope - Jillian

            self.ws.cell(row=2, column=1, value="Test Name").style = top_style
            self.ws.cell(row=2, column=2, value="Date").style = top_style
            self.ws.cell(row=2, column=3, value="Time").style = top_style
            self.ws.cell(row = 1, column = 4, value="Fan Test - V/I for power supplying all 4 fans, RD signal outputs").style = top_style
            self.ws.merge_cells(start_row=1, start_column=4, end_row=1, end_column=self.tc_res_first_col-1)
            
            self.ws.cell(row=2, column=4, value="Supply Voltage").style = top_style
            self.ws.cell(row=2, column=5, value="Supply Current").style = top_style

            for i in range(1,7):
                self.ws.cell(row=2, column=self.fan_rd_first_col-1+i, value=f"Fan {i} RD").style = top_style

            # Fan PWM sweep headers inserted right after fan_test() result columns
            self.ws.cell(
                row=1,
                column=self.fan_pwm_first_col,
                value="Fan PWM Sweep - Programmed/read fan voltage, current, and oscillation frequency"
            ).style = top_style

            self.ws.merge_cells(
                start_row=1,
                start_column=self.fan_pwm_first_col,
                end_row=1,
                end_column=self.fan_pwm_first_col + self.fan_pwm_cols - 1
            )

            for col_offset, header in enumerate(self.get_fan_pwm_headers()):
                self.ws.cell(
                    row=2,
                    column=self.fan_pwm_first_col + col_offset,
                    value=header
                ).style = top_style

            for i in range(1,5):    
                self.ws.cell(row=2, column=self.tc_res_first_col-1+i, value=f"TC{i}_Resistance").style = top_style
                self.ws.cell(row=2, column=self.tc_res_first_col-1+4+i, value=f"TC{i}_Temp_Rise").style = top_style

            self.ws.cell(row=1, column=self.tc_res_first_col, value="Heater Test - Results for each heating element and temperature rise after heating time").style = top_style
            self.ws.merge_cells(start_row=1, start_column=self.tc_res_first_col, end_row=1, end_column=self.hv_res_first_col-1)

            self.ws.cell(row=1, column=self.hv_res_first_col, value="HV Test - Results for each configuration. Resistance in units shown, time constant is tau (seconds) in a*e^(-tau * t)+c ").style = top_style
            self.ws.merge_cells(start_row=1, start_column=self.hv_res_first_col, end_row=1, end_column=self.hv_res_first_col+3+(7*self.hv_cols))
            for i in range(8):
                self.ws.cell(row=2, column=self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ Open Res").style = top_style
                self.ws.cell(row=2, column=1+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ Open Fit On").style = top_style
                self.ws.cell(row=2, column=2+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ Open Fit Off").style = top_style
                self.ws.cell(row=2, column=3+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ 10k Res").style = top_style
                self.ws.cell(row=2, column=4+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ 10k Fit On").style = top_style
                self.ws.cell(row=2, column=5+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}+ 10k Fit Off").style = top_style
                self.ws.cell(row=2, column=6+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- Open Res").style = top_style
                self.ws.cell(row=2, column=7+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- Open Fit On").style = top_style
                self.ws.cell(row=2, column=8+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- Open Fit Off").style = top_style
                self.ws.cell(row=2, column=9+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- 10k Res").style = top_style
                self.ws.cell(row=2, column=10+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- 10k Fit On").style = top_style
                self.ws.cell(row=2, column=11+self.hv_res_first_col+(i*self.hv_cols), value=f"Ch{i}- 10k Fit Off").style = top_style

            #Expands each column to have the best width to fit everything
            column_letters = tuple(openpyxl.utils.get_column_letter(col_number + 1) for col_number in range(self.ws.max_column))
            for column_letter in column_letters:
                self.ws.column_dimensions[column_letter].bestFit = True
            self.ws.freeze_panes = self.ws.cell(row=3, column=2)
            self.wb.save(self.path_to_spreadsheet)
            self.row = 3

        #In any case, start filling in the spreadsheet with initial test parameters we have right know
        #Test name starts in red so that if the test is incomplete it will show as a fail
        self.ws.cell(row=self.row, column=1, value=self.test_name).style = "fail"
        self.ws.cell(row=self.row, column=2, value=datetime.today().strftime('%m/%d/%Y'))
        self.ws.cell(row=self.row, column=3, value=datetime.today().strftime('%I:%M:%S %p'))
        self.wb.save(self.path_to_spreadsheet)
    def get_fan_pwm_headers(self):
        csv_headers = [
            "Programmed Fan Voltage [V]",
            "Read Fan Voltage [V]",
            "Read Fan Current [A]"
        ]

        for fan in range(1, self.fan_pwm_num_fans + 1):
            csv_headers.append(
                "Fan #" + str(fan) + " - Oscillation Frequency [Hz]"
            )

        return csv_headers


    def ensure_fan_pwm_columns(self):
        """
        Ensures the existing output_file has Fan PWM columns inserted between
        fan_test() results and Heater Test results.

        If the workbook was created before Fan PWM columns existed, this shifts
        the pre-existing Heater/HV columns to the right. Old rows will have blank
        Fan PWM entries.
        """

        fan_pwm_headers = self.get_fan_pwm_headers()

        # Check whether the Fan PWM section already exists.
        already_has_fan_pwm = (
            self.ws.cell(row=2, column=self.fan_pwm_first_col).value
            == fan_pwm_headers[0]
        )

        # Rebuild row-1 merged section headers safely.
        # This avoids stale merged-cell ranges after inserting columns.
        for merged_range in list(self.ws.merged_cells.ranges):
            if merged_range.min_row == 1 and merged_range.max_row == 1:
                self.ws.unmerge_cells(str(merged_range))

        if not already_has_fan_pwm:
            self.ws.insert_cols(self.fan_pwm_first_col, self.fan_pwm_cols)

        # Re-write top merged section headers using the new layout.
        self.ws.cell(
            row=1,
            column=4,
            value="Fan Test - V/I for power supplying all 4 fans, RD signal outputs"
        ).style = "top"

        self.ws.merge_cells(
            start_row=1,
            start_column=4,
            end_row=1,
            end_column=self.fan_pwm_first_col - 1
        )

        self.ws.cell(
            row=1,
            column=self.fan_pwm_first_col,
            value="Fan PWM Sweep - Programmed/read fan voltage, current, and oscillation frequency"
        ).style = "top"

        self.ws.merge_cells(
            start_row=1,
            start_column=self.fan_pwm_first_col,
            end_row=1,
            end_column=self.fan_pwm_first_col + self.fan_pwm_cols - 1
        )

        self.ws.cell(
            row=1,
            column=self.tc_res_first_col,
            value="Heater Test - Results for each heating element and temperature rise after heating time"
        ).style = "top"

        self.ws.merge_cells(
            start_row=1,
            start_column=self.tc_res_first_col,
            end_row=1,
            end_column=self.hv_res_first_col - 1
        )

        self.ws.cell(
            row=1,
            column=self.hv_res_first_col,
            value="HV Test - Results for each configuration. Resistance in units shown, time constant is tau (seconds) in a*e^(-tau * t)+c "
        ).style = "top"

        self.ws.merge_cells(
            start_row=1,
            start_column=self.hv_res_first_col,
            end_row=1,
            end_column=self.hv_res_first_col + 3 + (7 * self.hv_cols)
        )

        # Re-write Fan PWM column headers.
        for col_offset, header in enumerate(fan_pwm_headers):
            self.ws.cell(
                row=2,
                column=self.fan_pwm_first_col + col_offset,
                value=header
            ).style = "top"


    def write_fan_pwm_results_to_spreadsheet(self, fan_pwm_rows):
        """
        Writes Fan PWM sweep data into output_file.

        The first PWM sweep point is written on self.row, which is the same row
        as the nominal fan_test(), heater_test(), and hv_test() summary data.

        Additional sweep points are written on the following rows, only in the
        Fan PWM-specific columns.
        """

        # Make sure headers exist before writing data.
        for col_offset, header in enumerate(self.get_fan_pwm_headers()):
            self.ws.cell(
                row=2,
                column=self.fan_pwm_first_col + col_offset,
                value=header
            ).style = "top"

        for row_offset, fan_pwm_row in enumerate(fan_pwm_rows):
            excel_row = self.row + row_offset

            for col_offset, value in enumerate(fan_pwm_row):
                self.ws.cell(
                    row=excel_row,
                    column=self.fan_pwm_first_col + col_offset,
                    value=value
                )

        self.wb.save(self.path_to_spreadsheet)

    def fan_test(self, write_to_spreadsheet=True):
        #Fan test
        self.k.initialize_fan()
        self.r0.power("ON", "fan")
        self.r1.power("ON", "fanread")
        print(f"{self.prefix} --> Fans Tested at Nominal Voltage Specified ..... {self.json_data['rigol832a_fan_voltage']} ") # print out the "nominal voltage" aka the initial user specified fans voltage here, as it gets updated to max voltage after sweep test and becomes lost data
        print(f"{self.prefix} --> Fans turned on, waiting {self.json_data['fan_wait']} seconds for the fans to reach steady state...")
        time.sleep(self.json_data['fan_wait'])
        fan_voltage = self.r0.get_voltage("fan")
        fan_current = self.r0.get_current("fan")
        fanread_voltage = self.r1.get_voltage("fanread")
        fanread_current = self.r1.get_current("fanread")
        fan_read_signal = self.k.measure_fan()
        
        #take all data again for good measure
        fan_voltage = self.r0.get_voltage("fan")
        fan_current = self.r0.get_current("fan")
        fanread_voltage = self.r1.get_voltage("fanread")
        fanread_current = self.r1.get_current("fanread")
        fan_read_signal = self.k.measure_fan()
        
                
        self.r0.power("OFF", "fan")
        self.r1.power("OFF", "fanread")
        print(f"{self.prefix} --> Fans turned off")
        print(f"{self.prefix} --> Fan power supply was {fan_voltage}V and {fan_current}A")
        print(f"{self.prefix} --> Read signal for each fan was {fan_read_signal}")
        print(f"{self.prefix} --> Fan read pullup supply was {fanread_voltage}V and {fanread_current}A")
        
        self.datastore['fan_voltage'] = fan_voltage
        self.datastore['fan_current'] = fan_current
        self.datastore['fanread_voltage'] = fanread_voltage
        self.datastore['fanread_current'] = fanread_current
        self.datastore['fan_read_signal'] = fan_read_signal

        # During Fan PWM sweep, fan_test() is used only as a measurement helper.
        # Do not overwrite the nominal fan_test() spreadsheet results or
        # pass/fail status while sweeping through intentionally different voltages.
        if not write_to_spreadsheet:
            return None
        

        self.fan_test_result = True
        if ((fan_voltage < self.json_data["fan_voltage_max"]) and (fan_voltage > self.json_data["fan_voltage_min"])):
            self.ws.cell(row=self.row, column=4, value=fan_voltage)
            self.datastore['Tests']['fan_voltage_test'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=4, value=fan_voltage).style = "fail"
            self.datastore['Tests']['fan_voltage_test'] = "Fail"
            self.fan_test_result = False

        if ((fan_current < self.json_data["fan_current_max"]) and (fan_current > self.json_data["fan_current_min"])):
            self.ws.cell(row=self.row, column=5, value=fan_current)
            self.datastore['Tests']['fan_current_test'] = "Pass"
        else:
            self.ws.cell(row=self.row, column=5, value=fan_current).style = "fail"
            self.datastore['Tests']['fan_current_test'] = "Fail"
            self.fan_test_result = False

        for i in range(1,7):
            if ((fan_read_signal[i] < self.json_data["fan_read_max"]) and (fan_read_signal[i] > self.json_data["fan_read_min"])):
                self.ws.cell(row=self.row, column=self.fan_rd_first_col-1+i, value=round(fan_read_signal[i], self.rounding_factor))
                self.datastore['Tests'][f'fan_signal_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=self.fan_rd_first_col-1+i, value=round(fan_read_signal[i], self.rounding_factor)).style = "fail"
                self.fan_test_result = False

        return None
		# return fan_read_signal, fanread_voltage, fanread_current # return these for voltage sweep fan func plot

#---------------------------------------- Start of Fan PWM measurement ----------------------			
    def fan_test_sweep(self):
        #### all units for voltage in Volts here
        # ---- data storage lists
        # dictionaries where key = specific member (aka fan), value = y-axis variable (osc freq)
        num_fans = self.json_data['keysight970a_fan_num'] 
        fan_signals = {i: [] for i in range(1, num_fans + 1)} 
        prog_voltage = [] # program defined input fan voltage 
        read_voltage = [] # measured input fan voltage 
        read_current = [] # measured input fan current
        # create folder for storing fan data plots
        fan_results_folder = os.path.join(self.results_path, "fan PWM results") 
        os.makedirs(fan_results_folder, exist_ok=True) # check if folder exists, if not then create it

        # ----- start sweep test -----
		
        # initialize sweep parameters
        steps = int(self.json_data['rigol832a_fan_voltage_number_of_data_samples'])
        fan_voltage_max = self.json_data['rigol832a_fan_voltage_max']
        fan_voltage_min = self.json_data['rigol832a_fan_voltage_min']
        # check that step size is at least 2 (otherwise it is just a one voltage input value...)
        if (steps < 2) :
            print(f"{self.prefix} --> CONFIG ERROR! Invalid Number of Data Samples: ")			
            print(f"{self.prefix} --> Fan Sweep Test's number of data samples must be at least 2 to perform a sweep")	
            print(f"{self.prefix} --> Please fix rigol832a_fan_voltage_number_of_data_samples value in config.json and Restart test")	
            return None # end func execution early upon error
        
        # check that voltage max > voltage min input 
        if (fan_voltage_max <= fan_voltage_min) :
            print(f"{self.prefix} --> CONFIG ERROR! Invalid Fan Voltage Range: ")			
            print(f"{self.prefix} --> Fan Sweep Test's MAXIMUM voltage value is less than or equal to MINIMUM voltage value")	
            print(f"{self.prefix} --> Please fix rigol832a_fan_voltage_max and rigol832a_fan_voltage_min values in config.json and Restart test")	
            return None # end func execution early upon error
        
     
        # hardcode the sweep points' voltage values to ensure that max voltage value is max voltage specified if it isnt reached 
        for i in range(steps):
            voltage = fan_voltage_min + (fan_voltage_max - fan_voltage_min) * i / (steps - 1) # calculate voltage sweep point 
            prog_voltage.append(round(voltage, 3)) 
    
        # force the last sweep point to be fan_voltage max
        prog_voltage[-1] = fan_voltage_max # replace the last element with the specified max voltage input 
        print(f"{self.prefix} --> Fan Voltage Sweep Points in Volts : {prog_voltage}")
        
        
        # start sweeping
        for sweep_voltage in prog_voltage :
            self.json_data['rigol832a_fan_voltage'] = sweep_voltage
            self.r0.setup_fan() # apply new voltage to fan
            # call fan_test 
            self.fan_test(write_to_spreadsheet=False) # measurement only; do not overwrite nominal fan_test() row

            # extract info
            fan_read_signal = self.datastore['fan_read_signal']
            fan_voltage = self.datastore['fan_voltage']
            fan_current = self.datastore['fan_current']

            # parse info into collections
            read_voltage.append(fan_voltage)
            read_current.append(fan_current)

            for fan in range(1, num_fans + 1):
                fan_signals[fan].append( fan_read_signal.get(fan,-1) )

        # save info for greater spreadsheet storage
        self.datastore['fan_pwm_programmed_voltage'] = prog_voltage
        self.datastore['fan_pwm_read_voltage'] = read_voltage
        self.datastore['fan_pwm_read_current'] = read_current
        self.datastore['fan_pwm_signals'] = fan_signals
        # plot data (individual plots for each fan)
        # plot data: individual input voltage vs. frequency plot for each fan
        for fan in range(1, num_fans + 1):
            plt.figure()

            plt.plot( prog_voltage, fan_signals[fan], marker='o')

            plt.title( "Programmed Input Voltage vs. Fan #" + str(fan) + " PWM Frequency", fontsize=14)
            plt.xlabel("Programmed Fan Voltage [V]", fontsize=12)
            plt.ylabel("Oscillation Frequency [Hz]", fontsize=12)
            plt.grid(True)

            # store png plot in the folder
            plot_name = "fan_" + str(fan) + "_voltage_vs_frequency.png"
            save_path = os.path.join(fan_results_folder, plot_name)
            plt.savefig(save_path, bbox_inches="tight")
            plt.close()

        # plot data: programmed input voltage vs. read input voltage
        plt.figure()

        plt.plot(prog_voltage, read_voltage, marker='o')

        plt.title( "Programmed Fan Voltage vs. Read Fan Voltage", fontsize=14)
        plt.xlabel("Programmed Fan Voltage [V]", fontsize=12)
        plt.ylabel("Read Fan Voltage [V]", fontsize=12)
        plt.grid(True)

        # store png plot in the folder
        plot_name = "programmed_voltage_vs_read_voltage.png"
        save_path = os.path.join(fan_results_folder, plot_name)
        plt.savefig(save_path, bbox_inches="tight")
        plt.close()

        # build CSV-formatted Fan PWM table
        csv_headers = self.get_fan_pwm_headers()
        fan_pwm_rows = []

        for i in range(len(prog_voltage)):
            csv_row = [
                prog_voltage[i],
                read_voltage[i],
                read_current[i]
            ]

            for fan in range(1, num_fans + 1):
                csv_row.append(fan_signals[fan][i])

            fan_pwm_rows.append(csv_row)

        # save Fan PWM table into the existing Excel output_file
        # headers go in row 2; data goes in the current test row and rows below
        self.write_fan_pwm_results_to_spreadsheet(fan_pwm_rows)

        # still save standalone CSV copy in the fan PWM results folder
        csv_name = "fan_pwm_results.csv"
        csv_save_path = os.path.join(fan_results_folder, csv_name)

        with open(csv_save_path, mode="w", newline="") as csv_file:
            csv_writer = csv.writer(csv_file)

            csv_writer.writerow(csv_headers)
            csv_writer.writerows(fan_pwm_rows)

        print(f"{self.prefix} --> Fan PWM plots, CSV data, and Excel output_file data saved")
        print(f"{self.prefix} --> Fan PWM results folder: {fan_results_folder}")
        print(f"{self.prefix} --> Excel output_file: {self.path_to_spreadsheet}")
#---------------------------------------- End of Fan PWM measurement ----------------------		
	    
			
    def heater_test(self):
        #Heater test
        #First measure resistance of heating element with no power connected
        self.k.initialize_resistance()
        heater_resistance = self.k.measure_resistance()
        print(f"{self.prefix} --> Heating element resistances are {heater_resistance}")

        self.heat_test_result = True
        for i in range(1,5):
            if ((heater_resistance[i] < self.json_data["heating_element_max"]) and (heater_resistance[i] > self.json_data["heating_element_min"])):
                self.ws.cell(row=self.row, column=self.tc_res_first_col-1+i, value=round(heater_resistance[i], self.rounding_factor))
                self.datastore['Tests'][f'heating_element_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=self.tc_res_first_col-1+i, value=round(heater_resistance[i], self.rounding_factor)).style = "fail"
                self.datastore['Tests'][f'heating_element_test_{i}'] = "Fail"
                self.heat_test_result = False

        self.datastore['heater_resistance'] = heater_resistance

        #Then prepare RTD and switch relay to connect power
        self.k.initialize_rtd()
        temp1 = self.k.measure_rtd()
        #temp1 = self.k.measure_rtd() #measure again for good measure
        self.r0.power("ON", "heat_supply")
        self.r0.power("ON", "heat_switch")
        print(f"{self.prefix} --> Heat turned on, waiting {self.json_data['heat_wait']} seconds for the sensors to heat up...")
        time.sleep(self.json_data['heat_wait'])
        supply_voltage = self.r0.get_voltage("heat_supply")
        supply_current = self.r0.get_current("heat_supply")
        switch_voltage = self.r0.get_voltage("heat_switch")
        switch_current = self.r0.get_current("heat_switch")
        temp2 = self.k.measure_rtd()
        #temp2 = self.k.measure_rtd() #measure again for good measure
        temp_rise = []
        temp_rise.append(temp2[1] - temp1[1])
        temp_rise.append(temp2[2] - temp1[2])
        temp_rise.append(temp2[3] - temp1[3])
        temp_rise.append(temp2[4] - temp1[4])

        self.r0.power("OFF", "heat_supply")
        self.r0.power("OFF", "heat_switch")
        print(f"{self.prefix} --> Heat turned off")
        print(f"{self.prefix} --> Heat power supply was {supply_voltage}V and {supply_current}A")
        print(f"{self.prefix} --> Heat power switch was {switch_voltage}V and {switch_current}A")
        print(f"{self.prefix} --> Original temperatures were {temp1}")
        print(f"{self.prefix} --> Temperatures after {self.json_data['heat_wait']} seconds were {temp2}, a rise of {temp_rise}")

        for i in range(4):
            if ((temp_rise[i] < self.json_data["temp_increase_max"]) and (temp_rise[i] > self.json_data["temp_increase_min"])):
                self.ws.cell(row=self.row, column=self.tc_res_first_col+4+i, value=round(temp_rise[i], self.rounding_factor))
                self.datastore['Tests'][f'temperature_rise_test_{i}'] = "Pass"
            else:
                self.ws.cell(row=self.row, column=self.tc_res_first_col+4+i, value=round(temp_rise[i], self.rounding_factor)).style = "fail"
                self.datastore['Tests'][f'temperature_rise_test_{i}'] = "Fail"
                self.heat_test_result = False

        self.datastore['heater_supply_voltage'] = supply_voltage
        self.datastore['heater_supply_current'] = supply_current
        self.datastore['heater_switch_voltage'] = switch_voltage
        self.datastore['heater_switch_current'] = switch_current

        self.datastore['temp1'] = temp1
        self.datastore['temp2'] = temp2
        self.datastore['temp_rise'] = temp_rise

    def hv_test(self):
        #HV Leakage Test
        hv_results = {}
        self.r1.power("ON", "hvpullup")
        self.r1.power("ON", "hvpullup2")
        self.hv_test_result = True        
        # for i in self.json_data['channels_to_test']:
        if (self.json_data["simultaneous_test"] == "True"): #This distinction does not matter if only one channel total is being tested
            chs_to_test = self.json_data['channels_to_test'] #Test all channels at once
            total_chs_to_test = ["placeholder"]           #Only one full course of tests run - list contents not read
        else:
            total_chs_to_test = self.json_data['channels_to_test'] #Test channels one after the other
            chs_to_test = total_chs_to_test
                    
        for single_test in total_chs_to_test:
            single_test_done = False
            if (self.json_data["simultaneous_test"] == "True"):
                single_test = chs_to_test            	
            try:
                self.hv_test_single(single_test, hv_results)
                single_test_done = True
            except (ConnectionResetError, BrokenPipeError) as e:
                print(traceback.format_exc())
                print("Connection broken, attempting to reset...")
                self.c.turn_off(list(range(16)), emergency=True)             	    
                self.reset_pyvisa_connections()   
                self.r1.power("ON", "hvpullup")
                self.r1.power("ON", "hvpullup2")   
            except SystemExit as e:
                self.emergency_shutoff() 
                print(traceback.format_exc())
                print("Detecting exception",e,"but shutting off and continuing...")      
                self.r1.power("ON", "hvpullup")
                self.r1.power("ON", "hvpullup2")              	    	

            if not single_test_done:
                try:
                    self.hv_test_single(single_test, hv_results)  #try again 
                except: #give up
                    print("Detected exception, powering off all devices first.")
                    self.emergency_shutoff()       
                    raise          

        relay_done = False
        while not relay_done:
            try:
                self.r1.power("OFF", "hvpullup")
                self.r1.power("OFF", "hvpullup2")
                relay_done = True
            except (ConnectionResetError, BrokenPipeError, pyvisa.errors.VisaIOError) as e:
                print(traceback.format_exc())
                print("Connection broken, attempting to reset...")
                self.reset_pyvisa_connections()

        #Voltage is in volts, current is in microamps, R in Mohms
        for i in chs_to_test:
            try:
                hv_results[i]["pos_open_R"] = float(hv_results[i]["pos_open_V"])/float(hv_results[i]["pos_open_I"])
            except:
                hv_results[i]["pos_open_R"] = 0
            try:
                hv_results[i]["pos_term_R"] = float(hv_results[i]["pos_term_V"])/float(hv_results[i]["pos_term_I"])
            except Exception as e:
                hv_results[i]["pos_term_R"] = 0
                print(e)
                print(hv_results[i]["pos_term_V"])
                print(hv_results[i]["pos_term_I"])               
            try:
                hv_results[i]["neg_open_R"] = float(hv_results[i]["neg_open_V"])/float(hv_results[i]["neg_open_I"])
            except:
                hv_results[i]["neg_open_R"] = 0
            try:
                hv_results[i]["neg_term_R"] = float(hv_results[i]["neg_term_V"])/float(hv_results[i]["neg_term_I"])
            except Exception as e:            
                hv_results[i]["neg_term_R"] = 0
                print(e)  
                print(hv_results[i]["neg_term_V"])
                print(hv_results[i]["neg_term_I"])                
                

            print(f"{self.prefix} --> Channel {i} HV results are {hv_results[i]}")

            for num,j in enumerate(["pos_open_R", "neg_open_R"]):
                max_val = self.json_data["hv_resistance_open_max"]
                min_val = self.json_data["hv_resistance_open_min"]

                if ((float(hv_results[i][j]) < max_val) and (float(hv_results[i][j]) > min_val)):
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+(i*self.hv_cols)+(num*6), value=f"{round(float(hv_results[i][j]), self.rounding_factor)}Mohm")
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+(i*self.hv_cols)+(num*6), value=f"{round(float(hv_results[i][j]), self.rounding_factor)}Mohm").style = "fail"
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Fail"
                    self.hv_test_result = False

            for num,j in enumerate(["pos_term_R", "neg_term_R"]):
                max_val = self.json_data["hv_resistance_term_max"]
                min_val = self.json_data["hv_resistance_term_min"]

                if ((float(hv_results[i][j]) < max_val) and (float(hv_results[i][j]) > min_val)):
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+3+(i*self.hv_cols)+(num*6), value=f"{round(float(hv_results[i][j]*1E3), self.rounding_factor)}kohm")
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+3+(i*self.hv_cols)+(num*6), value=f"{round(float(hv_results[i][j]*1E3), self.rounding_factor)}kohm").style = "fail"
                    self.datastore['Tests'][f'hv_test_ch{i}_{j}'] = "Fail"
                    self.hv_test_result = False

            for num,j in enumerate(["pos_open", "pos_term", "neg_open", "neg_term"]):
                j_on = j + "_on_fit"
                j_off = j + "_off_fit"
                if ((float(hv_results[i][j_on][0][1]) < self.json_data["hv_tau_max"]) and (float(hv_results[i][j_on][0][1]) > self.json_data["hv_tau_min"])):
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+1+(i*self.hv_cols)+(num*3), value=round(float(hv_results[i][j_on][0][1]), self.rounding_factor))
                    self.datastore['Tests'][f'hv_on_fit_test_ch{i}_{j_on}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+1+(i*self.hv_cols)+(num*3), value=round(float(hv_results[i][j_on][0][1]), self.rounding_factor)).style = "fail"
                    self.datastore['Tests'][f'hv_on_fit_test_ch{i}_{j_on}'] = "Fail"
                    self.hv_test_result = False
                if ((float(hv_results[i][j_off][0][1]) < self.json_data["hv_tau_max"]) and (float(hv_results[i][j_off][0][1]) > self.json_data["hv_tau_min"])):
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+2+(i*self.hv_cols)+(num*3), value=round(float(hv_results[i][j_off][0][1]), self.rounding_factor))
                    self.datastore['Tests'][f'hv_off_fit_test_ch{i}_{j_off}'] = "Pass"
                else:
                    self.ws.cell(row=self.row, column=self.hv_res_first_col+2+(i*self.hv_cols)+(num*3), value=round(float(hv_results[i][j_off][0][1]), self.rounding_factor)).style = "fail"
                    self.datastore['Tests'][f'hv_off_fit_test_ch{i}_{j_off}'] = "Fail"
                    self.hv_test_result = False

            self.datastore[f'hv_ch{i}'] = {}
            for j in ["pos_open_V", "pos_open_I", "pos_open_R", "neg_open_V", "neg_open_I", "neg_open_R", "pos_open_on_fit", "pos_open_off_fit", "neg_open_on_fit", "neg_open_off_fit",
                      "pos_term_V", "pos_term_I", "pos_term_R", "neg_term_V", "neg_term_I", "neg_term_R", "pos_term_on_fit", "pos_term_off_fit", "neg_term_on_fit", "neg_term_off_fit"]:
                self.datastore[f'hv_ch{i}'][j] = hv_results[i][j]
        self.write_rc_calculated_vs_measured_summary()

    def hv_test_single(self, single_test, hv_results):
            if (self.json_data["simultaneous_test"] != "True"):
                chs_to_test = [single_test] #Test only one channel at a time
            else:
                chs_to_test = single_test #Test all channels (input is array)
            		
            #chs_to_test = self.json_data['channels_to_test']
            pos_chs = []
            neg_chs = []
            
            for i in chs_to_test:
                pos_chs.append(self.json_data[f"pcb_ch_{i}_pos"])
                neg_chs.append(self.json_data[f"pcb_ch_{i}_neg"])
                hv_results[i] = {}

            chs_string = ""
            for i in chs_to_test:
                chs_string = chs_string + str(i) + "_"

            #Measure the ramp from 0 to positive voltage with open termination
            v = self.json_data['caenR8033DM_open_voltage']
            relay_setting = 0
            for i in chs_to_test:
                relay_setting = relay_setting | (1 << i)  
            for pos_ch in pos_chs:
                self.c.set_HV_value(pos_ch, v)
                print(f"{self.prefix} --> Turning Channel {pos_ch} HV from 0 to {v}V with open termination")

            #self.k.set_relay(0, 1 << i) #<- how does this work?
            ramp_done = False
            while not ramp_done:
                try:
                    self.k.set_relay(0, relay_setting)
                    self.c.turn_on(pos_chs)
                    print(f"{self.prefix} --> HV reached max values, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
                    ramp_done = True
                    time.sleep(self.json_data['hv_stability_wait']) #may need to change for simul test?
                except (ConnectionResetError, BrokenPipeError) as e:
                    print(traceback.format_exc())
                    print("Connection broken, attempting to reset...")
                    self.c.turn_off(list(range(16)), emergency=True)
                    self.reset_pyvisa_connections()
                    self.r1.power("ON", "hvpullup")
                    self.r1.power("ON", "hvpullup2")



            csv_name = f"{self.test_name}_chs{chs_string}pos_open_on.csv"
            self.record_hv_data(csv_name)
            for i in chs_to_test:
                pos_ch = self.json_data[f"pcb_ch_{i}_pos"]
                fit = self.hv_curve_fit(csv_name, pos_ch, on = True, term = False)
                hv_results[i]["pos_open_on_fit"] = fit
                hv_results[i]["pos_open_V"] = self.c.get_voltage(pos_ch)
                hv_results[i]["pos_open_I"] = self.c.get_current(pos_ch)
                self.make_plot(csv_name, f"Ch {i} from 0 to {v}V, open termination", pos_ch, fit[0][1], [v-5, v+5])
                self.experimental_rc_charge_fit_all_signals(csv_name, pos_ch, i, "open_load", "pos")
            ###################

            #Measure the ramp from positive voltage to 0 with open termination
            for pos_ch in pos_chs:
                print(f"{self.prefix} --> Turning Channel {pos_ch} HV from {v}V to 0 with open termination")
            self.c.turn_off(pos_chs)
            # print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            # time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{chs_string}_pos_open_off.csv"
            self.record_hv_data(csv_name)
            for i in chs_to_test:
                pos_ch = self.json_data[f"pcb_ch_{i}_pos"]
                fit = self.hv_curve_fit(csv_name, pos_ch, on = False, term = False)
                hv_results[i]["pos_open_off_fit"] = fit
                self.make_plot(csv_name, f"Ch {i} from {v} to 0V, open termination", pos_ch, fit[0][1])
                self.experimental_rc_discharge_fit_all_signals(csv_name, pos_ch, i, "open_load", "pos")
            ###################

            #Measure the ramp from 0 to positive voltage with 10k termination
            v = self.json_data['caenR8033DM_term_voltage']
            for pos_ch in pos_chs:
                self.c.set_HV_value(pos_ch, v)
                print(f"{self.prefix} --> Turning Channel {pos_ch} HV from 0 to {v}V with 10k termination")
            ramp_done = False
            while not ramp_done:
                try:
                    self.k.set_relay(0, 0)
                    self.c.turn_on(pos_chs)
                    ramp_done = True
                    print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_termination_wait']} seconds to stabilize...")
                    time.sleep(self.json_data['hv_termination_wait'])
                except (ConnectionResetError, BrokenPipeError) as e:
                    print(traceback.format_exc())
                    print("Connection broken, attempting to reset...")
                    self.c.turn_off(list(range(16)), emergency=True)
                    self.reset_pyvisa_connections()
                    self.r1.power("ON", "hvpullup")
                    self.r1.power("ON", "hvpullup2")


            csv_name = f"{self.test_name}_ch{chs_string}_pos_term_on.csv"
            self.record_hv_data(csv_name, short_time=True)
            for i in chs_to_test:
                pos_ch = self.json_data[f"pcb_ch_{i}_pos"]
                fit = self.hv_curve_fit(csv_name, pos_ch, on = True, term = True)
                hv_results[i]["pos_term_on_fit"] = fit
                #hv_results[i]["pos_term_V"] = self.c.get_voltage(pos_ch)
                #hv_results[i]["pos_term_I"] = self.c.get_current(pos_ch)


                print("Measuring 10k terminated voltage and current...")
                voltage = self.c.get_voltage(pos_ch, print_meas=True)
                current = self.c.get_current(pos_ch, print_meas=True)

                #print(f"Ch {i} voltage: {voltage}, current {current}, resistance {voltage/current}")
                hv_results[i]["pos_term_V"] = voltage
                hv_results[i]["pos_term_I"] = current

                self.make_plot(csv_name, f"Ch {i} from 0 to {v}V, termination resistor", pos_ch, fit[0][1], [v-5, v+5])
                self.experimental_rc_charge_fit_all_signals(csv_name, pos_ch, i, "termination_10k", "pos")
            time.sleep(5)
            ###################

            #Measure the ramp from positive voltage to 0 with 10k termination
            for pos_ch in pos_chs:
                print(f"{self.prefix} --> Turning Channel {pos_ch} HV from {v}V to 0 with 10k termination")
            self.c.turn_off(pos_chs)
            # print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            # time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{chs_string}_pos_term_off.csv"
            self.record_hv_data(csv_name, short_time=True)
            for i in chs_to_test:
                pos_ch = self.json_data[f"pcb_ch_{i}_pos"]
                fit = self.hv_curve_fit(csv_name, pos_ch, on = False, term = True)
                hv_results[i]["pos_term_off_fit"] = fit
                self.make_plot(csv_name, f"Ch {i} from {v} to 0V, termination resistor", pos_ch, fit[0][1])
                self.experimental_rc_discharge_fit_all_signals(csv_name, pos_ch, i, "termination_10k", "pos")
            ###################

            #Measure the ramp from 0 to negative voltage with open termination
            v = self.json_data['caenR8033DM_open_voltage']
            for neg_ch in neg_chs:
                self.c.set_HV_value(neg_ch, v)
                print(f"{self.prefix} --> Turning Channel {neg_ch} HV from 0 to -{v}V with open termination")
            ramp_done = False
            while not ramp_done:
                try:
                    self.k.set_relay(relay_setting, relay_setting)
                    self.c.turn_on(neg_chs)
                    ramp_done = True
                    print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
                    time.sleep(self.json_data['hv_stability_wait'])
                except (ConnectionResetError, BrokenPipeError) as e:
                    print(traceback.format_exc())
                    print("Connection broken, attempting to reset...")
                    self.c.turn_off(list(range(16)), emergency=True)
                    self.reset_pyvisa_connections()
                    self.r1.power("ON", "hvpullup")
                    self.r1.power("ON", "hvpullup2")

            csv_name = f"{self.test_name}_ch{chs_string}_neg_open_on.csv"
            self.record_hv_data(csv_name)
            for i in chs_to_test:
                neg_ch = self.json_data[f"pcb_ch_{i}_neg"]
                fit = self.hv_curve_fit(csv_name, neg_ch, on = True, term = False)
                hv_results[i]["neg_open_on_fit"] = fit
                hv_results[i]["neg_open_V"] = self.c.get_voltage(neg_ch)
                hv_results[i]["neg_open_I"] = self.c.get_current(neg_ch)
                self.make_plot(csv_name, f"Ch {i} from 0 to -{v}V, open termination", neg_ch, fit[0][1], [v-5, v+5])
                self.experimental_rc_charge_fit_all_signals(csv_name, neg_ch, i, "open_load", "neg")
            ###################

            #Measure the ramp from negative voltage to 0 with open termination
            for neg_ch in neg_chs:
                print(f"{self.prefix} --> Turning Channel {neg_ch} HV from -{v}V to 0 with open termination")
            self.c.turn_off(neg_chs)
            # print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            # time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{chs_string}_neg_open_off.csv"
            self.record_hv_data(csv_name)
            for i in chs_to_test:
                neg_ch = self.json_data[f"pcb_ch_{i}_neg"]
                fit = self.hv_curve_fit(csv_name, neg_ch, on = False, term = False)
                hv_results[i]["neg_open_off_fit"] = fit
                self.make_plot(csv_name, f"Ch {i} from -{v} to 0V, open termination", neg_ch, fit[0][1])
                self.experimental_rc_discharge_fit_all_signals(csv_name, neg_ch, i, "open_load", "neg")
            ###################
            
            #Measure the ramp from 0 to negative voltage with 10k termination
            v = self.json_data['caenR8033DM_term_voltage']
            for neg_ch in neg_chs:
                self.c.set_HV_value(neg_ch, v)
                print(f"{self.prefix} --> Turning Channel {neg_ch} HV from 0 to -{v}V with 10k termination")
            ramp_done = False

            while not ramp_done:
                try:
                    self.k.set_relay(relay_setting, 0)
                    self.c.turn_on(neg_chs)
                    ramp_done = True
                    print(f"{self.prefix} --> HV reached max value, waiting {self.json_data['hv_termination_wait']} seconds to stabilize...")
                    time.sleep(self.json_data['hv_termination_wait'])
                except (ConnectionResetError, BrokenPipeError) as e:
                    print(traceback.format_exc())
                    print("Connection broken, attempting to reset...")
                    self.c.turn_off(list(range(16)), emergency=True)
                    self.reset_pyvisa_connections()
                    self.r1.power("ON", "hvpullup")
                    self.r1.power("ON", "hvpullup2")
                # except ValueError as e:
                    # print(e)
                    # #print("Positive channel voltages:", self.c.get_current(pos_chs))

                    # self.c.turn_off(neg_chs)

                    # pullups_ok = False
                    # while not pullups_ok:
                        # try:
                            # hvv = self.r1.get_voltage("hvpullup")
                            # hv2v = self.r1.get_voltage("hvpullup2")
                            # if hvv >= 18 and hv2v >= 18: #check if they are on
                                # pullups_ok = True
                            # else:
                                # print("Resetting and power cycling pullup")
                                # print(hvv, hv2v)
                                # raise ValueError("")
                        # except Exception as e:
                            # print(e)
                            # try:
                                # self.reset_pullup_rigol_and_power_on()
                                # time.sleep(5)
                            # except Exception as e:
                                # time.sleep(5)
                                # pass #run again
                    # # self.reset_pyvisa_connections()
                    # # self.r1.power("ON", "hvpullup")
                    # # self.r1.power("ON", "hvpullup2")
                    # # time.sleep(1)
                    # print("Resetting relays")
                    # #self.k.keysight.close()
                    # #self.k = Keysight970A(self.rm, self.json_data)
                    # #time.sleep(1)
                    # self.k.set_relay(0, relay_setting)
                    # print("Trying to fire relay again")
                    # #try block will run again

            csv_name = f"{self.test_name}_ch{chs_string}_neg_term_on.csv"
            self.record_hv_data(csv_name, short_time=True)
            for i in chs_to_test:
                neg_ch = self.json_data[f"pcb_ch_{i}_neg"]          
                fit = self.hv_curve_fit(csv_name, neg_ch, on = True, term = True)
                hv_results[i]["neg_term_on_fit"] = fit
                #hv_results[i]["neg_term_V"] = self.c.get_voltage(neg_ch)
                #hv_results[i]["neg_term_I"] = self.c.get_current(neg_ch)
                print("Measuring 10k terminated voltage and current...")
                voltage = self.c.get_voltage(neg_ch)#, print_meas=True)
                voltage = self.c.get_voltage(neg_ch)
                current = self.c.get_current(neg_ch)#, print_meas=True)
                current = self.c.get_current(neg_ch)
                #print(f"Ch {i} voltage: {voltage}, current {current}, resistance {voltage/current}")
                hv_results[i]["neg_term_V"] = voltage
                hv_results[i]["neg_term_I"] = current
                self.make_plot(csv_name, f"Ch {i} from 0 to -{v}V, termination resistor", neg_ch, fit[0][1], [v-5, v+5])
                self.experimental_rc_charge_fit_all_signals(csv_name, neg_ch, i, "termination_10k", "neg")                
            time.sleep(5) #debug
            ###################            
            
            #Measure the ramp from 0 to negative voltage with 10k termination
            for neg_ch in neg_chs:
                print(f"{self.prefix} --> Turning Channel {neg_ch} HV from -{v}V to 0 with 10k termination")
            self.c.turn_off(neg_chs)
            # print(f"{self.prefix} --> HV turned off, waiting {self.json_data['hv_stability_wait']} seconds to stabilize...")
            # time.sleep(self.json_data['hv_stability_wait'])

            csv_name = f"{self.test_name}_ch{chs_string}_neg_term_off.csv"
            self.record_hv_data(csv_name, short_time=True)
            for i in chs_to_test:
                neg_ch = self.json_data[f"pcb_ch_{i}_neg"]              
                fit = self.hv_curve_fit(csv_name, neg_ch, on = False, term = True)
                hv_results[i]["neg_term_off_fit"] = fit
                self.make_plot(csv_name, f"Ch {i} from -{v} to 0V, termination resistor", neg_ch, fit[0][1])
                self.experimental_rc_discharge_fit_all_signals(csv_name, neg_ch, i, "termination_10k", "neg")
            ###################            


            
    def emergency_shutoff(self):
        self.c.turn_off(list(range(16)), emergency=True) #Turn off HV channels
        #input("pause here")
        self.r0.power("OFF", "heat_supply") #Turn off fan and heater power
        self.r0.power("OFF", "heat_switch")
        self.r0.power("OFF", "fan")
        self.r1.power("OFF", "fanread")
        self.k.set_relay(0, 0) #Probably not necessary    

    def reset_pyvisa_connections(self):	
        self.r0.rigol.close()
        self.r0 = RigolDP832A(self.rm, self.json_data, 0)
        self.r0.setup_fan()
        self.r0.setup_heater_supply()
        self.r0.setup_heater_switch()
            
        self.r1.rigol.close()
        self.r1 = RigolDP832A(self.rm, self.json_data, 1)
        self.r1.setup_hvpullup()
        self.r1.setup_hvpullup2()
        self.r1.setup_fanread()  
            	
        self.k.keysight.close()
        self.k = Keysight970A(self.rm, self.json_data)    
    	
    def record_hv_data(self, name, short_time=False):
        data = []
        cycle_start_time = time.time()
        time_string = datetime.fromtimestamp(cycle_start_time)
        prev_measurement = cycle_start_time - 1
        if short_time:
            minutes_wait = self.json_data['hv_minutes_duration_short']
        else:
            minutes_wait = self.json_data['hv_minutes_duration_long']
        print(f"{self.prefix} --> Collecting data for {name} for {minutes_wait} minutes starting at {time_string}...")
        while (time.time() - cycle_start_time < (minutes_wait * 60)):
            if (time.time() > prev_measurement + self.json_data['hv_seconds_interval']):
                #print(f"{self.prefix} --> Measurement taken at {time.time()}")
                prev_measurement = prev_measurement + self.json_data['hv_seconds_interval']
                datum = [datetime.now()]
                for i in range(16):
                    datum.append(self.c.get_voltage(i))
                    datum.append(self.c.get_current(i))
                #datum.append(self.c.get_voltage(list(range(16))))
                data.append(datum)
        with open(os.path.join(self.results_path, name), 'w') as fp:
            csv_writer = csv.writer(fp, delimiter=',')
            csv_writer.writerows(data)
        #input("ok?")

    def hv_curve_fit(self, name, ch, on = True, term = False):
        ch_datetime = []
        ch_voltage = []
        ch_current = []
        with open(os.path.join(self.results_path, name), 'r', newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            for row in spamreader:
                ch_datetime.append(datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S.%f'))
                ch_voltage.append(float(row[1 + (ch*2)]))
                if (term):
                    ch_current.append(float(row[2 + (ch*2)])/1000)
                else:
                    ch_current.append(float(row[2 + (ch*2)]))
        if (on):
            data = ch_current
        else:
            data = ch_voltage
        #print(f"For channel {ch}, grabbed column {2 + (ch*2)} and got this data")
        #print(ch_current)
        first_time = ch_datetime[0]
        ch_timedelta = [i-first_time for i in ch_datetime]
        ch_time = [datetime(2024, 1, 1, 0, i.seconds//60%60, i.seconds%60, 0) for i in ch_timedelta]

        def exp_fit(x, a, b, c):
            y = a*np.exp(-b*x) + c
            return y

        first_timestamp = ch_time[0].timestamp()
        time_seconds = [dt.timestamp() - first_timestamp for dt in ch_time]
        try:
            fit = curve_fit(exp_fit, time_seconds, data)
        except RuntimeError:
            fit = [[0,0]]
        #The result is an array like
        #[
        #    "[-0.0003239   0.04760632  0.29665177]",

        #    "[[ 4.03901010e-10 -5.76474012e-08 -2.84934772e-12],
        #    [-5.76474012e-08  2.38199865e-05 -7.01897955e-09],
        #    [-2.84934772e-12 -7.01897955e-09  1.26525944e-11]]"
        #]
        #The first array is the convergant results for the 3 parameters a, b, and c
        #The second array is the confidence levels for each based on the covarience with the other variables
        #Low numbers less than one mean that the confidence is high
        #https://docs.scipy.org/doc/scipy/reference/generated/scipy.optimize.curve_fit.html
        return fit

    
    def make_hv_plots(self): #not used
        ch0_pos_open_fit = self.datastore['hv_ch0']['pos_open_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_pos_open_on", "Ch {i} from 0 to 2kV, open termination", True, True, 0, ch0_pos_open_fit)
        # self.make_plot(f"{self.test_name}_ch0_pos_open_off", "2kV to 0, open termination", False, True)
        ch0_pos_term_fit = self.datastore['hv_ch0']['pos_term_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_pos_10k_on", "Ch {i} from 0 to 2kV, 10k termination", True, True, 0, ch0_pos_term_fit)
        #self.make_plot(f"{self.test_name}_ch0_pos_10k_off", "2kV to 0, 10k termination", False, True, 0)

        ch0_neg_open_fit = self.datastore['hv_ch0']['neg_open_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_neg_open_on", "Ch {i} from 0 to -2kV, open termination", True, False, 8, ch0_neg_open_fit)
        # self.make_plot(f"{self.test_name}_ch0_neg_open_off", "-2kV to 0, open termination", False, False)
        ch0_neg_term_fit = self.datastore['hv_ch0']['neg_term_fit'][0][1]
        self.make_plot(f"{self.test_name}_ch0_neg_10k_on", "Ch {i} from 0 to -2kV, 10k termination", True, False, 8, ch0_neg_term_fit)
        #self.make_plot(f"{self.test_name}_ch0_neg_10k_off", "-2kV to 0, 10k termination", False, True, 8)

    def make_plot(self, filename, name, ch, fit=None, axes = None):
        ch1_time, ch1_voltage, ch1_current = self.get_ch_data(os.path.join(self.results_path, filename), ch)
        # self.make_plot(f"{self.test_name}_ch0_neg_10k_off", "-2kV to 0, 10k termination", False, False)

        fig = plt.figure(figsize=(16, 12), dpi=80)
        ax = fig.add_subplot(1,1,1)

        ax.plot(ch1_time, ch1_current, label="Ch Current")
        self.format_plot(ax)
                                           
                                   
                                             

        ax2 = ax.twinx()
                                                                  
                      
        ax2.plot(ch1_time, ch1_voltage, label="Ch Voltage", color="red")
                                                                                                             

        fig.suptitle((name), fontsize=36)
                                

        ax.set_xlabel("Time (Minutes:Seconds)", fontsize=24)
        ax.set_ylabel("Current (uA)", fontsize=24)

        # ax.set_xlim([0,150])
        if (axes):
            ax2.set_ylim([axes[0],axes[1]])
        ax2.set_ylabel("Voltage (V)", fontsize=24)
        self.format_plot(ax2)

        if fit:
            textstr = r'$\tau=%.4f$' % (fit)
            props = dict(boxstyle='round', facecolor='wheat', alpha=0.5)
            ax.text(0.75, 0.75, textstr, transform=ax.transAxes, fontsize=24,
            verticalalignment='top', bbox=props)

        fig.legend(loc='lower left', prop={'size': 20}, ncol=2)
        stem = Path(filename).stem
        fig.savefig(os.path.join(self.results_path, f"{stem}_ch{ch}.png"))
        plt.close(fig)

    def get_ch_data(self, data_file, ch):
        ch1_datetime = []
        ch1_voltage = []
        ch1_current = []
        with open(data_file, 'r', newline='') as csvfile:
            spamreader = csv.reader(csvfile, delimiter=',')
            for row in spamreader:
                ch1_datetime.append(datetime.strptime(row[0], '%Y-%m-%d %H:%M:%S.%f'))
                ch1_voltage.append(float(row[1+(ch*2)]))
                ch1_current.append(float(row[2+(ch*2)]))

        first_time = ch1_datetime[0]
        ch1_timedelta = [i-first_time for i in ch1_datetime]
        ch1_time = [datetime(2024, 1, 1, 0, i.seconds//60%60, i.seconds%60, 0) for i in ch1_timedelta]
        return ch1_time, ch1_voltage, ch1_current

    def format_plot(self, ax):
        tick_size = 18
        ax.xaxis.set_major_formatter(mdates.DateFormatter('%M:%S'))
        ax.tick_params(axis='x', labelsize=tick_size, colors='black')  # Set tick size and color here
        ax.tick_params(axis='y', labelsize=tick_size, colors='black')  # Set tick size and color here
   
   
   #------------------------------ true RC model functions for the 2 pole filter ------------------------------
    def get_rc_filter_components(self, hv_ch):
        """
        Returns RC component values for one internal HV channel.

        Internal channels 0-3 use rc_filter_ch0_3_r_ohm and rc_filter_ch0_3_c_f.
        Internal channels 4-7 use rc_filter_ch4_7_r_ohm and rc_filter_ch4_7_c_f.
        """

        hv_ch = int(hv_ch)

        if hv_ch < 0 or hv_ch > 7:
            raise ValueError(f"hv_ch must be internal channel number 0 through 7. Got {hv_ch}.")

        if hv_ch <= 3:
            r_value = float(self.json_data["rc_filter_ch0_3_r_ohm"])
            c_value = float(self.json_data["rc_filter_ch0_3_c_f"])
            bank = "ch0_3"
        else:
            r_value = float(self.json_data["rc_filter_ch4_7_r_ohm"])
            c_value = float(self.json_data["rc_filter_ch4_7_c_f"])
            bank = "ch4_7"

        if r_value <= 0:
            raise ValueError(f"RC filter resistance for channel {hv_ch} must be > 0.")

        if c_value <= 0:
            raise ValueError(f"RC filter capacitance for channel {hv_ch} must be > 0.")

        return r_value, r_value, c_value, c_value, bank

    def get_rc_filter_tolerances(self, hv_ch):
        """
        Returns R and C tolerance percentages for one internal HV channel.

        Internal channels 0-3 use rc_filter_ch0_3_*_tolerance_percent.
        Internal channels 4-7 use rc_filter_ch4_7_*_tolerance_percent.
        """

        hv_ch = int(hv_ch)

        if hv_ch < 0 or hv_ch > 7:
            raise ValueError(f"hv_ch must be internal channel number 0 through 7. Got {hv_ch}.")

        if hv_ch <= 3:
            r_tolerance_percent = float(self.json_data.get("rc_filter_ch0_3_r_tolerance_percent", 0.0))
            c_tolerance_percent = float(self.json_data.get("rc_filter_ch0_3_c_tolerance_percent", 0.0))
        else:
            r_tolerance_percent = float(self.json_data.get("rc_filter_ch4_7_r_tolerance_percent", 0.0))
            c_tolerance_percent = float(self.json_data.get("rc_filter_ch4_7_c_tolerance_percent", 0.0))

        return abs(r_tolerance_percent), abs(c_tolerance_percent)


    def two_pole_rc_time_constants(self, r_top, r_bottom, c_top, c_bottom, load_ohm=None):
        """
        Analytically calculates the two natural time constants of a directly
        cascaded 2-pole RC ladder. This does not use scipy curve_fit.
        """

        g_top = 1.0 / r_top
        g_bottom = 1.0 / r_bottom
        g_load = 0.0 if load_ohm is None else 1.0 / float(load_ohm)

        if load_ohm is not None and float(load_ohm) <= 0:
            raise ValueError("load_ohm must be positive, or None for open load.")

        a11 = -(g_top + g_bottom) / c_top
        a12 = g_bottom / c_top
        a21 = g_bottom / c_bottom
        a22 = -(g_bottom + g_load) / c_bottom

        trace = a11 + a22
        determinant = (a11 * a22) - (a12 * a21)
        discriminant = (trace * trace) - (4.0 * determinant)

        if discriminant < 0.0 and abs(discriminant) < 1e-18:
            discriminant = 0.0

        if discriminant < 0.0:
            raise ValueError("RC pole calculation produced complex poles. Check component values and topology assumptions.")

        sqrt_disc = discriminant ** 0.5
        pole_1 = 0.5 * (trace + sqrt_disc)
        pole_2 = 0.5 * (trace - sqrt_disc)

        if pole_1 >= 0.0 or pole_2 >= 0.0:
            raise ValueError("RC pole calculation produced a non-decaying pole. Check component values and load model.")

        tau_1 = -1.0 / pole_1
        tau_2 = -1.0 / pole_2
        tau_fast = min(tau_1, tau_2)
        tau_slow = max(tau_1, tau_2)

        return {"tau_fast_s": float(tau_fast), "tau_slow_s": float(tau_slow), "dominant_tau_s": float(tau_slow), "settling_5tau_s": float(5.0 * tau_slow), "pole_fast_1_per_s": float(-1.0 / tau_fast), "pole_slow_1_per_s": float(-1.0 / tau_slow)}

    def two_pole_rc_time_constants_with_tolerance(self, r_top, r_bottom, c_top, c_bottom, r_tolerance_percent, c_tolerance_percent, load_ohm=None):
        """
        Calculates nominal, minimum, and maximum RC time constants using R/C tolerances.

        The min/max range is found by checking all component tolerance corners:
            R_top min/max
            R_bottom min/max
            C_top min/max
            C_bottom min/max

        This is better than assuming tau = R*C because the circuit is a loaded
        two-pole ladder, especially for the 10k termination case.
        """

        nominal_result = self.two_pole_rc_time_constants(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, load_ohm=load_ohm)

        r_tol = abs(float(r_tolerance_percent)) / 100.0
        c_tol = abs(float(c_tolerance_percent)) / 100.0

        r_top_values = [max(r_top * (1.0 - r_tol), 1e-30), r_top * (1.0 + r_tol)]
        r_bottom_values = [max(r_bottom * (1.0 - r_tol), 1e-30), r_bottom * (1.0 + r_tol)]
        c_top_values = [max(c_top * (1.0 - c_tol), 1e-30), c_top * (1.0 + c_tol)]
        c_bottom_values = [max(c_bottom * (1.0 - c_tol), 1e-30), c_bottom * (1.0 + c_tol)]

        tau_fast_values = []
        tau_slow_values = []
        dominant_tau_values = []
        settling_5tau_values = []

        for r_top_corner in r_top_values:
            for r_bottom_corner in r_bottom_values:
                for c_top_corner in c_top_values:
                    for c_bottom_corner in c_bottom_values:
                        corner_result = self.two_pole_rc_time_constants(r_top=r_top_corner, r_bottom=r_bottom_corner, c_top=c_top_corner, c_bottom=c_bottom_corner, load_ohm=load_ohm)
                        tau_fast_values.append(corner_result["tau_fast_s"])
                        tau_slow_values.append(corner_result["tau_slow_s"])
                        dominant_tau_values.append(corner_result["dominant_tau_s"])
                        settling_5tau_values.append(corner_result["settling_5tau_s"])

        nominal_result["r_tolerance_percent"] = float(r_tolerance_percent)
        nominal_result["c_tolerance_percent"] = float(c_tolerance_percent)

        nominal_result["tau_fast_min_s"] = float(min(tau_fast_values))
        nominal_result["tau_fast_max_s"] = float(max(tau_fast_values))

        nominal_result["tau_slow_min_s"] = float(min(tau_slow_values))
        nominal_result["tau_slow_max_s"] = float(max(tau_slow_values))

        nominal_result["dominant_tau_min_s"] = float(min(dominant_tau_values))
        nominal_result["dominant_tau_max_s"] = float(max(dominant_tau_values))

        nominal_result["settling_5tau_min_s"] = float(min(settling_5tau_values))
        nominal_result["settling_5tau_max_s"] = float(max(settling_5tau_values))

        return nominal_result
    
    def two_pole_rc_state_matrix(self, r_top, r_bottom, c_top, c_bottom, load_ohm=None):
        """
        Builds the 2-state RC ladder matrix.
        State vector: x = [node1_voltage, output_voltage]
        """

        g_top = 1.0 / r_top
        g_bottom = 1.0 / r_bottom
        g_load = 0.0 if load_ohm is None else 1.0 / float(load_ohm)

        return np.array([[-(g_top + g_bottom) / c_top, g_bottom / c_top], [g_bottom / c_bottom, -(g_bottom + g_load) / c_bottom]], dtype=float)


    def two_pole_rc_step_curve(self, r_top, r_bottom, c_top, c_bottom, source_voltage, load_ohm=None, n_points=500, time_multiplier=5.0, transition="step_on"):
        """
        Generates modeled voltage/current data points for a 2-pole RC filter.

        transition="step_on": source steps from 0 V to source_voltage.
        transition="step_off": circuit starts charged, then source steps to 0 V.
        """

        n_points = max(2, int(n_points))
        time_multiplier = max(1.0, float(time_multiplier))
        source_voltage = float(source_voltage)

        tau_info = self.two_pole_rc_time_constants(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, load_ohm=load_ohm)
        t_end = max(time_multiplier * tau_info["dominant_tau_s"], 1e-12)
        time_s = np.linspace(0.0, t_end, n_points)

        A = self.two_pole_rc_state_matrix(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, load_ohm=load_ohm)
        input_on = np.array([source_voltage / (r_top * c_top), 0.0], dtype=float)
        steady_on = -np.linalg.solve(A, input_on)

        if transition == "step_on":
            x_initial = np.array([0.0, 0.0], dtype=float)
            x_final = steady_on
            source_after_switch = source_voltage
        elif transition == "step_off":
            x_initial = steady_on
            x_final = np.array([0.0, 0.0], dtype=float)
            source_after_switch = 0.0
        else:
            raise ValueError(f"transition must be 'step_on' or 'step_off'. Got {transition}.")

        eigvals, eigvecs = np.linalg.eig(A)
        coeff = np.linalg.solve(eigvecs, x_initial - x_final)
        modes = coeff[:, None] * np.exp(eigvals[:, None] * time_s[None, :])
        state = (x_final[:, None] + eigvecs @ modes).T
        state = np.real_if_close(state, tol=1000)

        if np.iscomplexobj(state):
            raise ValueError("RC curve generation produced complex state values. Check RC values and load assumptions.")

        state = state.astype(float)
        node1_voltage = state[:, 0]
        output_voltage = state[:, 1]
        source_voltage_array = np.full_like(time_s, source_after_switch, dtype=float)
        source_current = (source_voltage_array - node1_voltage) / r_top
        load_current = np.zeros_like(time_s, dtype=float) if load_ohm is None else output_voltage / float(load_ohm)

        return {"time_s": time_s, "source_voltage_v": source_voltage_array, "node1_voltage_v": node1_voltage, "output_voltage_v": output_voltage, "source_current_a": source_current, "load_current_a": load_current}


    def save_rc_curve_csv(self, curve_data, csv_path):
        """
        Saves modeled RC voltage/current curve points to CSV.
        """

        with open(csv_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Time [s]", "Source Voltage [V]", "Node 1 Voltage [V]", "Output Voltage [V]", "Source Current [A]", "Load Current [A]"])

            for i in range(len(curve_data["time_s"])):
                writer.writerow([curve_data["time_s"][i], curve_data["source_voltage_v"][i], curve_data["node1_voltage_v"][i], curve_data["output_voltage_v"][i], curve_data["source_current_a"][i], curve_data["load_current_a"][i]])


    def save_rc_curve_plot(self, curve_data, plot_path, title):
        """
        Saves one RC model plot with voltage and current versus time.
        """

        fig = plt.figure(figsize=(14, 9), dpi=100)
        ax1 = fig.add_subplot(1, 1, 1)

        ax1.plot(curve_data["time_s"], curve_data["node1_voltage_v"], label="Node 1 Voltage")
        ax1.plot(curve_data["time_s"], curve_data["output_voltage_v"], label="Output Voltage")
        ax1.set_xlabel("Time [s]", fontsize=14)
        ax1.set_ylabel("Voltage [V]", fontsize=14)
        ax1.grid(True)

        ax2 = ax1.twinx()
        ax2.plot(curve_data["time_s"], curve_data["source_current_a"], label="Source Current")

        if np.any(np.abs(curve_data["load_current_a"]) > 0.0):
            ax2.plot(curve_data["time_s"], curve_data["load_current_a"], label="Load Current")

        ax2.set_ylabel("Current [A]", fontsize=14)

        lines_1, labels_1 = ax1.get_legend_handles_labels()
        lines_2, labels_2 = ax2.get_legend_handles_labels()
        fig.legend(lines_1 + lines_2, labels_1 + labels_2, loc="lower left", fontsize=12)

        fig.suptitle(title, fontsize=18)
        fig.tight_layout()
        fig.savefig(plot_path, bbox_inches="tight")
        plt.close(fig)

    def save_theoretical_rc_bank_summary_plots(self, rc_results_folder, curve_points, time_multiplier):
        """
        Saves one theoretical summary plot for channels 0-3 and one for channels 4-7.
        """

        self.save_theoretical_rc_bank_summary_plot(bank_start=0, bank_end=3, bank_label="ch0_3", rc_results_folder=rc_results_folder, curve_points=curve_points, time_multiplier=time_multiplier)
        self.save_theoretical_rc_bank_summary_plot(bank_start=4, bank_end=7, bank_label="ch4_7", rc_results_folder=rc_results_folder, curve_points=curve_points, time_multiplier=time_multiplier)


    def save_theoretical_rc_bank_summary_plot(self, bank_start, bank_end, bank_label, rc_results_folder, curve_points, time_multiplier):
        """
        Saves a theoretical normalized output-voltage plot for one RC component bank.

        Theoretical time constants are polarity-independent, so this plot is shared by:
            0 -> +Vdd
            +Vdd -> 0
            0 -> -Vss
            -Vss -> 0
        """

        hv_ch = int(bank_start)
        r_top, r_bottom, c_top, c_bottom, bank = self.get_rc_filter_components(hv_ch)
        r_tolerance_percent, c_tolerance_percent = self.get_rc_filter_tolerances(hv_ch)

        termination_ohm = float(self.json_data.get("rc_filter_termination_ohm", 10000.0))

        open_tau = self.two_pole_rc_time_constants_with_tolerance(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, r_tolerance_percent=r_tolerance_percent, c_tolerance_percent=c_tolerance_percent, load_ohm=None)
        term_tau = self.two_pole_rc_time_constants_with_tolerance(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, r_tolerance_percent=r_tolerance_percent, c_tolerance_percent=c_tolerance_percent, load_ohm=termination_ohm)

        open_curve = self.two_pole_rc_step_curve(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, source_voltage=1.0, load_ohm=None, n_points=curve_points, time_multiplier=time_multiplier, transition="step_on")
        term_curve = self.two_pole_rc_step_curve(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, source_voltage=1.0, load_ohm=termination_ohm, n_points=curve_points, time_multiplier=time_multiplier, transition="step_on")

        plot_path = os.path.join(rc_results_folder, f"theoretical_{bank_label}_rc_filter_summary.png")

        fig = plt.figure(figsize=(14, 9), dpi=100)
        ax = fig.add_subplot(1, 1, 1)

        ax.plot(open_curve["time_s"], open_curve["output_voltage_v"], label="Open load output / source")
        ax.plot(term_curve["time_s"], term_curve["output_voltage_v"], label="10k termination output / source")

        ax.set_title(f"Theoretical RC Filter Model - Channels {bank_start}-{bank_end}", fontsize=18)
        ax.set_xlabel("Time [s]", fontsize=14)
        ax.set_ylabel("Normalized Output Voltage [V/V]", fontsize=14)
        ax.grid(True)
        ax.legend(loc="lower right", fontsize=11)

        textstr = (
            f"Component bank: {bank_label}\n"
            f"R_top = R_bottom = {r_top:.4g} Ohm ± {r_tolerance_percent:.4g}%\n"
            f"C_top = C_bottom = {c_top:.4g} F ± {c_tolerance_percent:.4g}%\n\n"
            f"Open dominant tau:\n"
            f"  nominal = {open_tau['dominant_tau_s']:.4g} s\n"
            f"  min = {open_tau['dominant_tau_min_s']:.4g} s\n"
            f"  max = {open_tau['dominant_tau_max_s']:.4g} s\n\n"
            f"10k dominant tau:\n"
            f"  nominal = {term_tau['dominant_tau_s']:.4g} s\n"
            f"  min = {term_tau['dominant_tau_min_s']:.4g} s\n"
            f"  max = {term_tau['dominant_tau_max_s']:.4g} s"
        )

        props = dict(boxstyle="round", facecolor="wheat", alpha=0.5)
        ax.text(0.03, 0.97, textstr, transform=ax.transAxes, fontsize=11, verticalalignment="top", bbox=props)

        fig.tight_layout()
        fig.savefig(plot_path, bbox_inches="tight")
        plt.close(fig)

        return plot_path
            
#------- experimental RC fit functions that use the above model as a basis for fitting real data -------
    def get_rc_results_folder(self):
        if hasattr(self, "rc_filter_results_folder"):
            return self.rc_filter_results_folder

        rc_results_folder = os.path.join(self.results_path, "RC filter model results")
        os.makedirs(rc_results_folder, exist_ok=True)
        self.rc_filter_results_folder = rc_results_folder

        return rc_results_folder


    def read_hv_csv_for_rc_fit(self, csv_name, hv_supply_ch):
        data_file = csv_name if os.path.isabs(csv_name) else os.path.join(self.results_path, csv_name)
        ch_datetime = []
        ch_voltage = []
        ch_current = []

        with open(data_file, "r", newline="") as csvfile:
            spamreader = csv.reader(csvfile, delimiter=",")

            for row in spamreader:
                if len(row) < 3 + (hv_supply_ch * 2):
                    continue

                ch_datetime.append(datetime.strptime(row[0], "%Y-%m-%d %H:%M:%S.%f"))
                ch_voltage.append(float(row[1 + (hv_supply_ch * 2)]))
                ch_current.append(float(row[2 + (hv_supply_ch * 2)]))

        if len(ch_datetime) < 4:
            raise ValueError(f"Not enough data points in {csv_name} for experimental RC fit.")

        first_time = ch_datetime[0]
        time_s = np.array([(dt - first_time).total_seconds() for dt in ch_datetime], dtype=float)
        voltage_v = np.array(ch_voltage, dtype=float)
        current_ua = np.array(ch_current, dtype=float)

        finite_mask = np.isfinite(time_s) & np.isfinite(voltage_v) & np.isfinite(current_ua)
        return {"time_s": time_s[finite_mask], "voltage_v": voltage_v[finite_mask], "current_ua": current_ua[finite_mask], "source_csv": data_file}


    def estimate_single_tau_guess(self, time_s, y_data, transition):
        y0 = float(y_data[0])
        yf = float(y_data[-1])
        t_end = max(float(time_s[-1]), 1e-9)

        if abs(yf - y0) < 1e-30:
            return max(t_end / 5.0, 1e-9)

        if transition == "charge":
            target = y0 + 0.6321205588 * (yf - y0)
        else:
            target = yf + 0.3678794412 * (y0 - yf)

        idx = int(np.argmin(np.abs(y_data - target)))

        if idx <= 0:
            return max(t_end / 5.0, 1e-9)

        return max(float(time_s[idx]), 1e-9)


    def single_exp_fallback_fit(self, time_s, y_data, transition, error_message):
        y0 = float(y_data[0])
        yf = float(y_data[-1])
        tau_guess = self.estimate_single_tau_guess(time_s, y_data, transition)

        if transition == "charge":
            fit_y = y0 + (yf - y0) * (1.0 - np.exp(-time_s / tau_guess))
        else:
            fit_y = yf + (y0 - yf) * np.exp(-time_s / tau_guess)

        ss_res = float(np.sum((y_data - fit_y) ** 2))
        ss_tot = float(np.sum((y_data - np.mean(y_data)) ** 2))
        r_squared = 0.0 if ss_tot <= 0.0 else 1.0 - (ss_res / ss_tot)

        return {"fit_ok": False, "fit_type": "single_exponential_fallback", "tau_fast_s": float(tau_guess), "tau_slow_s": float(tau_guess), "dominant_tau_s": float(tau_guess), "r_squared": float(r_squared), "fit_error": str(error_message), "fit_y": fit_y}


    def fit_rc_charge_curve(self, time_s, y_data):
        def charge_model(t, y0, a_fast, a_slow, tau_fast, tau_slow):
            return y0 + a_fast * (1.0 - np.exp(-t / tau_fast)) + a_slow * (1.0 - np.exp(-t / tau_slow))

        try:
            time_s = np.array(time_s, dtype=float)
            y_data = np.array(y_data, dtype=float)
            tau_guess = self.estimate_single_tau_guess(time_s, y_data, "charge")
            y_min = float(np.min(y_data))
            y_max = float(np.max(y_data))
            y_span = max(abs(y_max - y_min), 1e-12)
            t_end = max(float(time_s[-1]), 1e-9)
            positive_dt = np.diff(time_s)
            positive_dt = positive_dt[positive_dt > 0.0]
            tau_min = max(float(np.min(positive_dt)) / 100.0, 1e-9) if len(positive_dt) > 0 else 1e-9
            tau_max = max(t_end * 100.0, tau_min * 1000.0)
            delta = float(y_data[-1] - y_data[0])
            p0 = [float(y_data[0]), 0.35 * delta, 0.65 * delta, max(tau_guess / 5.0, tau_min), max(tau_guess, tau_min * 10.0)]
            lower_bounds = [y_min - 2.0 * y_span, -10.0 * y_span, -10.0 * y_span, tau_min, tau_min]
            upper_bounds = [y_max + 2.0 * y_span, 10.0 * y_span, 10.0 * y_span, tau_max, tau_max]

            popt, pcov = curve_fit(charge_model, time_s, y_data, p0=p0, bounds=(lower_bounds, upper_bounds), maxfev=20000)

            fit_y = charge_model(time_s, *popt)
            tau_fast = min(float(popt[3]), float(popt[4]))
            tau_slow = max(float(popt[3]), float(popt[4]))
            ss_res = float(np.sum((y_data - fit_y) ** 2))
            ss_tot = float(np.sum((y_data - np.mean(y_data)) ** 2))
            r_squared = 0.0 if ss_tot <= 0.0 else 1.0 - (ss_res / ss_tot)

            return {"fit_ok": True, "fit_type": "two_exponential_charge", "tau_fast_s": tau_fast, "tau_slow_s": tau_slow, "dominant_tau_s": tau_slow, "r_squared": float(r_squared), "fit_error": "", "fit_y": fit_y}

        except Exception as e:
            return self.single_exp_fallback_fit(time_s, y_data, "charge", e)


    def fit_rc_discharge_curve(self, time_s, y_data):
        def discharge_model(t, y_final, a_fast, a_slow, tau_fast, tau_slow):
            return y_final + a_fast * np.exp(-t / tau_fast) + a_slow * np.exp(-t / tau_slow)

        try:
            time_s = np.array(time_s, dtype=float)
            y_data = np.array(y_data, dtype=float)
            tau_guess = self.estimate_single_tau_guess(time_s, y_data, "discharge")
            y_min = float(np.min(y_data))
            y_max = float(np.max(y_data))
            y_span = max(abs(y_max - y_min), 1e-12)
            t_end = max(float(time_s[-1]), 1e-9)
            positive_dt = np.diff(time_s)
            positive_dt = positive_dt[positive_dt > 0.0]
            tau_min = max(float(np.min(positive_dt)) / 100.0, 1e-9) if len(positive_dt) > 0 else 1e-9
            tau_max = max(t_end * 100.0, tau_min * 1000.0)
            delta = float(y_data[0] - y_data[-1])
            p0 = [float(y_data[-1]), 0.35 * delta, 0.65 * delta, max(tau_guess / 5.0, tau_min), max(tau_guess, tau_min * 10.0)]
            lower_bounds = [y_min - 2.0 * y_span, -10.0 * y_span, -10.0 * y_span, tau_min, tau_min]
            upper_bounds = [y_max + 2.0 * y_span, 10.0 * y_span, 10.0 * y_span, tau_max, tau_max]

            popt, pcov = curve_fit(discharge_model, time_s, y_data, p0=p0, bounds=(lower_bounds, upper_bounds), maxfev=20000)

            fit_y = discharge_model(time_s, *popt)
            tau_fast = min(float(popt[3]), float(popt[4]))
            tau_slow = max(float(popt[3]), float(popt[4]))
            ss_res = float(np.sum((y_data - fit_y) ** 2))
            ss_tot = float(np.sum((y_data - np.mean(y_data)) ** 2))
            r_squared = 0.0 if ss_tot <= 0.0 else 1.0 - (ss_res / ss_tot)

            return {"fit_ok": True, "fit_type": "two_exponential_discharge", "tau_fast_s": tau_fast, "tau_slow_s": tau_slow, "dominant_tau_s": tau_slow, "r_squared": float(r_squared), "fit_error": "", "fit_y": fit_y}

        except Exception as e:
            return self.single_exp_fallback_fit(time_s, y_data, "discharge", e)


    def save_experimental_rc_fit_outputs(self, measurement_data, fit_result, hv_ch, polarity, load_case, transition, signal_key):
        rc_results_folder = self.get_rc_results_folder()
        file_stem = f"measured_ch{hv_ch}_{polarity}_{load_case}_{transition}_{signal_key}"
        fit_csv_path = os.path.join(rc_results_folder, f"{file_stem}_fit_data.csv")
        fit_plot_path = os.path.join(rc_results_folder, f"{file_stem}_fit_plot.png")

        measured_signal = measurement_data["voltage_v"] if signal_key == "voltage" else measurement_data["current_ua"]
        signal_label = "Voltage [V]" if signal_key == "voltage" else "Current [uA]"

        with open(fit_csv_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["Time [s]", "Measured Voltage [V]", "Measured Current [uA]", f"Fit Signal - {signal_label}", f"Fit Curve - {signal_label}"])

            for i in range(len(measurement_data["time_s"])):
                writer.writerow([measurement_data["time_s"][i], measurement_data["voltage_v"][i], measurement_data["current_ua"][i], measured_signal[i], fit_result["fit_y"][i]])

        fig = plt.figure(figsize=(14, 9), dpi=100)
        ax1 = fig.add_subplot(1, 1, 1)
        ax1.plot(measurement_data["time_s"], measurement_data["voltage_v"], label="Measured Voltage")

        if signal_key == "voltage":
            ax1.plot(measurement_data["time_s"], fit_result["fit_y"], label="Voltage Fit")

        ax1.set_xlabel("Time [s]", fontsize=14)
        ax1.set_ylabel("Voltage [V]", fontsize=14)
        ax1.grid(True)

        ax2 = ax1.twinx()
        ax2.plot(measurement_data["time_s"], measurement_data["current_ua"], label="Measured Current")

        if signal_key == "current":
            ax2.plot(measurement_data["time_s"], fit_result["fit_y"], label="Current Fit")

        ax2.set_ylabel("Current [uA]", fontsize=14)

        lines_1, labels_1 = ax1.get_legend_handles_labels()
        lines_2, labels_2 = ax2.get_legend_handles_labels()
        fig.legend(lines_1 + lines_2, labels_1 + labels_2, loc="lower left", fontsize=12)

        try:
            r_top, r_bottom, c_top, c_bottom, bank = self.get_rc_filter_components(hv_ch)
            r_tolerance_percent, c_tolerance_percent = self.get_rc_filter_tolerances(hv_ch)
        except:
            r_top, r_bottom, c_top, c_bottom, bank = 0.0, 0.0, 0.0, 0.0, "unknown"
            r_tolerance_percent, c_tolerance_percent = 0.0, 0.0

        calculated_tau = self._get_calculated_rc_tau(hv_ch, load_case)
        calculated_tau_min = self._get_calculated_rc_tau_min(hv_ch, load_case)
        calculated_tau_max = self._get_calculated_rc_tau_max(hv_ch, load_case)

        calc_text = "calculated tau unavailable"
        if calculated_tau is not None and calculated_tau_min is not None and calculated_tau_max is not None:
            calc_text = f"calc tau = {calculated_tau:.4g} s\ncalc range = [{calculated_tau_min:.4g}, {calculated_tau_max:.4g}] s"

        textstr = (
            f"Measured fit signal: {signal_key}\n"
            f"fit type = {fit_result['fit_type']}\n"
            f"tau_fast = {fit_result['tau_fast_s']:.4g} s\n"
            f"tau_slow = {fit_result['tau_slow_s']:.4g} s\n"
            f"dominant tau = {fit_result['dominant_tau_s']:.4g} s\n"
            f"R² = {fit_result['r_squared']:.4f}\n\n"
            f"R = {r_top:.4g} Ohm ± {r_tolerance_percent:.4g}%\n"
            f"C = {c_top:.4g} F ± {c_tolerance_percent:.4g}%\n"
            f"{calc_text}"
        )

        props = dict(boxstyle="round", facecolor="wheat", alpha=0.5)
        ax1.text(0.58, 0.97, textstr, transform=ax1.transAxes, fontsize=10, verticalalignment="top", bbox=props)

        fig.suptitle(f"Measured RC Fit - Ch {hv_ch}, {polarity}, {load_case}, {transition}", fontsize=18)
        fig.tight_layout()
        fig.savefig(fit_plot_path, bbox_inches="tight")

        return fit_csv_path, fit_plot_path



    def record_experimental_rc_result(self, hv_ch, polarity, load_case, transition, signal_key, fit_result, fit_csv_path, fit_plot_path, source_csv):
        if "rc_filter_experimental" not in self.datastore:
            self.datastore["rc_filter_experimental"] = {}

        if hv_ch not in self.datastore["rc_filter_experimental"]:
            self.datastore["rc_filter_experimental"][hv_ch] = {}

        if load_case not in self.datastore["rc_filter_experimental"][hv_ch]:
            self.datastore["rc_filter_experimental"][hv_ch][load_case] = {}

        if transition not in self.datastore["rc_filter_experimental"][hv_ch][load_case]:
            self.datastore["rc_filter_experimental"][hv_ch][load_case][transition] = {}

        if polarity not in self.datastore["rc_filter_experimental"][hv_ch][load_case][transition]:
            self.datastore["rc_filter_experimental"][hv_ch][load_case][transition][polarity] = {}

        self.datastore["rc_filter_experimental"][hv_ch][load_case][transition][polarity][signal_key] = {
            "signal_key": signal_key,
            "fit_ok": fit_result["fit_ok"],
            "fit_type": fit_result["fit_type"],
            "tau_fast_s": fit_result["tau_fast_s"],
            "tau_slow_s": fit_result["tau_slow_s"],
            "dominant_tau_s": fit_result["dominant_tau_s"],
            "r_squared": fit_result["r_squared"],
            "fit_error": fit_result["fit_error"],
            "source_csv": source_csv,
            "fit_csv": fit_csv_path,
            "fit_plot_png": fit_plot_path
        }

    def get_experimental_rc_signal_keys(self):
        """
        Returns which measured signals should be fit experimentally.
        Default is both voltage and current.
        """

        signal_keys = self.json_data.get("rc_filter_summary_signal_keys", ["voltage", "current"])

        if isinstance(signal_keys, str):
            signal_keys = [signal_keys]

        valid_signal_keys = []

        for signal_key in signal_keys:
            if signal_key in ["voltage", "current"]:
                valid_signal_keys.append(signal_key)

        if len(valid_signal_keys) == 0:
            valid_signal_keys = ["voltage", "current"]

        return valid_signal_keys


    def experimental_rc_charge_fit_all_signals(self, csv_name, hv_supply_ch, hv_ch, load_case, polarity):
        results = {}

        for signal_key in self.get_experimental_rc_signal_keys():
            results[signal_key] = self.experimental_rc_charge_fit(csv_name, hv_supply_ch, hv_ch, load_case, polarity, signal_key=signal_key)

        return results


    def experimental_rc_discharge_fit_all_signals(self, csv_name, hv_supply_ch, hv_ch, load_case, polarity):
        results = {}

        for signal_key in self.get_experimental_rc_signal_keys():
            results[signal_key] = self.experimental_rc_discharge_fit(csv_name, hv_supply_ch, hv_ch, load_case, polarity, signal_key=signal_key)

        return results
            
    def experimental_rc_charge_fit(self, csv_name, hv_supply_ch, hv_ch, load_case, polarity, signal_key="voltage"):
        try:
            measurement_data = self.read_hv_csv_for_rc_fit(csv_name, hv_supply_ch)
            measured_signal = measurement_data["voltage_v"] if signal_key == "voltage" else measurement_data["current_ua"]
            fit_result = self.fit_rc_charge_curve(measurement_data["time_s"], measured_signal)
            fit_csv_path, fit_plot_path = self.save_experimental_rc_fit_outputs(measurement_data, fit_result, hv_ch, polarity, load_case, "charge", signal_key)
            self.record_experimental_rc_result(hv_ch, polarity, load_case, "charge", signal_key, fit_result, fit_csv_path, fit_plot_path, measurement_data["source_csv"])
            return fit_result

        except Exception as e:
            print(f"{self.prefix} --> Experimental RC charge fit failed for Ch {hv_ch}, {polarity}, {load_case}: {e}")
            return {"fit_ok": False, "fit_type": "failed", "tau_fast_s": 0.0, "tau_slow_s": 0.0, "dominant_tau_s": 0.0, "r_squared": 0.0, "fit_error": str(e)}


    def experimental_rc_discharge_fit(self, csv_name, hv_supply_ch, hv_ch, load_case, polarity, signal_key="voltage"):
        try:
            measurement_data = self.read_hv_csv_for_rc_fit(csv_name, hv_supply_ch)
            measured_signal = measurement_data["voltage_v"] if signal_key == "voltage" else measurement_data["current_ua"]
            fit_result = self.fit_rc_discharge_curve(measurement_data["time_s"], measured_signal)
            fit_csv_path, fit_plot_path = self.save_experimental_rc_fit_outputs(measurement_data, fit_result, hv_ch, polarity, load_case, "discharge", signal_key)
            self.record_experimental_rc_result(hv_ch, polarity, load_case, "discharge", signal_key, fit_result, fit_csv_path, fit_plot_path, measurement_data["source_csv"])
            return fit_result

        except Exception as e:
            print(f"{self.prefix} --> Experimental RC discharge fit failed for Ch {hv_ch}, {polarity}, {load_case}: {e}")
            return {"fit_ok": False, "fit_type": "failed", "tau_fast_s": 0.0, "tau_slow_s": 0.0, "dominant_tau_s": 0.0, "r_squared": 0.0, "fit_error": str(e)}

    def store_rc_filter_time_constants(self):
        """
        Calculates, plots, and stores modeled RC filter data before any hardware tests.

        Outputs are written to:
            self.results_path / "RC filter model results"

        Files created:
            - rc_filter_time_constants_summary.csv
            - modeled curve CSV files
            - matching PNG plots
        """

        rc_results_folder = os.path.join(self.results_path, "RC filter model results")
        os.makedirs(rc_results_folder, exist_ok=True)

        termination_ohm = float(self.json_data.get("rc_filter_termination_ohm", 10000.0))
        curve_points = int(self.json_data.get("rc_filter_curve_points", 500))
        time_multiplier = float(self.json_data.get("rc_filter_curve_time_multiplier", 5.0))

        open_source_voltage = abs(float(self.json_data.get("rc_filter_open_model_voltage", self.json_data.get("caenR8033DM_open_voltage", 1.0))))
        term_source_voltage = abs(float(self.json_data.get("rc_filter_term_model_voltage", self.json_data.get("caenR8033DM_term_voltage", open_source_voltage))))

        summary_csv_path = os.path.join(rc_results_folder, "rc_filter_time_constants_summary.csv")

        self.datastore["rc_filter_model"] = {"results_folder": rc_results_folder, "summary_csv": summary_csv_path, "curve_points": curve_points, "curve_time_multiplier": time_multiplier, "open_load": {}, "termination_10k": {}}
        self.rc_filter_results_folder = rc_results_folder
        self.datastore["rc_filter_experimental"] = {}

        summary_rows = []
        load_cases = [{"case_name": "open_load", "load_label": "open", "load_ohm": None, "source_voltage": open_source_voltage}, {"case_name": "termination_10k", "load_label": f"{termination_ohm} ohm", "load_ohm": termination_ohm, "source_voltage": term_source_voltage}]
        transitions = ["step_on", "step_off"]

        for hv_ch in range(8):
            r_top, r_bottom, c_top, c_bottom, bank = self.get_rc_filter_components(hv_ch)

            for load_case in load_cases:
                case_name = load_case["case_name"]
                load_label = load_case["load_label"]
                load_ohm = load_case["load_ohm"]
                source_voltage = load_case["source_voltage"]

                r_tolerance_percent, c_tolerance_percent = self.get_rc_filter_tolerances(hv_ch)
                rc_result = self.two_pole_rc_time_constants_with_tolerance(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, r_tolerance_percent=r_tolerance_percent, c_tolerance_percent=c_tolerance_percent, load_ohm=load_ohm)
                
                rc_result["load"] = load_label
                rc_result["component_bank"] = bank
                rc_result["r_top_ohm"] = r_top
                rc_result["r_bottom_ohm"] = r_bottom
                rc_result["c_top_f"] = c_top
                rc_result["c_bottom_f"] = c_bottom
                rc_result["r_tolerance_percent"] = r_tolerance_percent
                rc_result["c_tolerance_percent"] = c_tolerance_percent
                rc_result["source_voltage_v"] = source_voltage
                rc_result["load_ohm"] = "open" if load_ohm is None else load_ohm

                for transition in transitions:
                    curve_data = self.two_pole_rc_step_curve(r_top=r_top, r_bottom=r_bottom, c_top=c_top, c_bottom=c_bottom, source_voltage=source_voltage, load_ohm=load_ohm, n_points=curve_points, time_multiplier=time_multiplier, transition=transition)
                    file_stem = f"ch{hv_ch}_{case_name}_{transition}"
                    curve_csv_path = os.path.join(rc_results_folder, f"{file_stem}_curve.csv")
                    plot_path = os.path.join(rc_results_folder, f"{file_stem}_voltage_current.png")
                    plot_title = f"RC Filter Model - Ch {hv_ch}, {load_label}, {transition.replace('_', ' ')}"

                    self.save_rc_curve_csv(curve_data, curve_csv_path)
                    self.save_rc_curve_plot(curve_data, plot_path, plot_title)

                    rc_result[f"{transition}_curve_csv"] = curve_csv_path
                    rc_result[f"{transition}_plot_png"] = plot_path

                    summary_rows.append([hv_ch, case_name, transition, bank, r_top, r_bottom, c_top, 
                                         c_bottom, r_tolerance_percent, c_tolerance_percent, 
                                         source_voltage, "open" if load_ohm is None else load_ohm, 
                                         rc_result["tau_fast_s"], rc_result["tau_fast_min_s"], 
                                         rc_result["tau_fast_max_s"], rc_result["tau_slow_s"], 
                                         rc_result["tau_slow_min_s"], rc_result["tau_slow_max_s"], 
                                         rc_result["dominant_tau_s"], rc_result["dominant_tau_min_s"], 
                                         rc_result["dominant_tau_max_s"], rc_result["settling_5tau_s"], 
                                         rc_result["settling_5tau_min_s"], rc_result["settling_5tau_max_s"], 
                                         rc_result["pole_fast_1_per_s"], rc_result["pole_slow_1_per_s"], 
                                         curve_csv_path, plot_path])
                self.datastore["rc_filter_model"][case_name][hv_ch] = rc_result

        with open(summary_csv_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(["HV Channel", "Load Case", "Transition", "Component Bank", "R Top [Ohm]", "R Bottom [Ohm]", "C Top [F]", "C Bottom [F]", "R Tolerance [%]", "C Tolerance [%]", "Source Voltage [V]", "Load Resistance [Ohm]", "Tau Fast Nominal [s]", "Tau Fast Min [s]", "Tau Fast Max [s]", "Tau Slow Nominal [s]", "Tau Slow Min [s]", "Tau Slow Max [s]", "Dominant Tau Nominal [s]", "Dominant Tau Min [s]", "Dominant Tau Max [s]", "5 Tau Settling Nominal [s]", "5 Tau Settling Min [s]", "5 Tau Settling Max [s]", "Pole Fast [1/s]", "Pole Slow [1/s]", "Curve CSV Path", "Plot PNG Path"])
            writer.writerows(summary_rows)

        self.save_theoretical_rc_bank_summary_plots(rc_results_folder, curve_points, time_multiplier)

        print(f"{self.prefix} --> RC filter model calculated before all hardware tests")
        print(f"{self.prefix} --> RC filter model results folder: {rc_results_folder}")
        print(f"{self.prefix} --> RC filter constants summary CSV: {summary_csv_path}")

#----------------------------- functions for comparing experimental RC fit results to calculated model -----------------------------
#----------------------------- functions for comparing experimental RC fit results to calculated model -----------------------------
    def _get_measured_rc_tau_for_polarity(self, hv_ch, load_case, transition, polarity, signal_key):
        """
        Returns the measured dominant tau for one exact case:
            channel + load + charge/discharge + pos/neg + voltage/current

        No positive/negative averaging is performed.
        """

        try:
            polarity_result = self.datastore["rc_filter_experimental"][hv_ch][load_case][transition][polarity]
        except:
            return None

        if isinstance(polarity_result, dict) and signal_key in polarity_result:
            result = polarity_result[signal_key]
        else:
            return None

        tau = float(result.get("dominant_tau_s", 0.0))

        if tau > 0.0 and np.isfinite(tau):
            return tau

        return None


    def _get_calculated_rc_tau(self, hv_ch, load_case):
        try:
            return float(self.datastore["rc_filter_model"][load_case][hv_ch]["dominant_tau_s"])
        except:
            return None


    def _get_calculated_rc_tau_min(self, hv_ch, load_case):
        try:
            return float(self.datastore["rc_filter_model"][load_case][hv_ch]["dominant_tau_min_s"])
        except:
            return None


    def _get_calculated_rc_tau_max(self, hv_ch, load_case):
        try:
            return float(self.datastore["rc_filter_model"][load_case][hv_ch]["dominant_tau_max_s"])
        except:
            return None


    def write_rc_calculated_vs_measured_summary(self):
        rc_results_folder = self.get_rc_results_folder()
        comparison_csv_path = os.path.join(rc_results_folder, "rc_filter_calculated_vs_measured_summary.csv")

        headers = ["Metric"]
        column_specs = []

        for hv_ch in range(8):
            for load_case, load_label in [("open_load", "Open"), ("termination_10k", "10k")]:
                for polarity, polarity_label in [("pos", "Pos"), ("neg", "Neg")]:
                    for transition, transition_label in [("charge", "Charge"), ("discharge", "Discharge")]:
                        headers.append(f"Ch{hv_ch} {load_label} {polarity_label} {transition_label} [s]")
                        column_specs.append((hv_ch, load_case, polarity, transition))

        calculated_nominal_row = ["calculated nominal"]
        calculated_min_row = ["calculated min"]
        calculated_max_row = ["calculated max"]

        measured_voltage_row = ["measured voltage"]
        measured_current_row = ["measured current"]

        percent_difference_voltage_row = ["percent difference voltage from nominal"]
        percent_difference_current_row = ["percent difference current from nominal"]

        voltage_tolerance_check_row = ["voltage measured inside calculated tolerance range"]
        current_tolerance_check_row = ["current measured inside calculated tolerance range"]

        for hv_ch, load_case, polarity, transition in column_specs:
            calculated_tau = self._get_calculated_rc_tau(hv_ch, load_case)
            calculated_tau_min = self._get_calculated_rc_tau_min(hv_ch, load_case)
            calculated_tau_max = self._get_calculated_rc_tau_max(hv_ch, load_case)

            measured_voltage_tau = self._get_measured_rc_tau_for_polarity(hv_ch, load_case, transition, polarity, "voltage")
            measured_current_tau = self._get_measured_rc_tau_for_polarity(hv_ch, load_case, transition, polarity, "current")

            calculated_nominal_row.append("" if calculated_tau is None else calculated_tau)
            calculated_min_row.append("" if calculated_tau_min is None else calculated_tau_min)
            calculated_max_row.append("" if calculated_tau_max is None else calculated_tau_max)

            measured_voltage_row.append("" if measured_voltage_tau is None else measured_voltage_tau)
            measured_current_row.append("" if measured_current_tau is None else measured_current_tau)

            if calculated_tau is None or measured_voltage_tau is None or calculated_tau == 0.0:
                percent_difference_voltage_row.append("")
            else:
                percent_difference_voltage_row.append(100.0 * (measured_voltage_tau - calculated_tau) / calculated_tau)

            if calculated_tau is None or measured_current_tau is None or calculated_tau == 0.0:
                percent_difference_current_row.append("")
            else:
                percent_difference_current_row.append(100.0 * (measured_current_tau - calculated_tau) / calculated_tau)

            if calculated_tau_min is None or calculated_tau_max is None or measured_voltage_tau is None:
                voltage_tolerance_check_row.append("")
            elif measured_voltage_tau >= calculated_tau_min and measured_voltage_tau <= calculated_tau_max:
                voltage_tolerance_check_row.append("PASS")
            else:
                voltage_tolerance_check_row.append("FAIL")

            if calculated_tau_min is None or calculated_tau_max is None or measured_current_tau is None:
                current_tolerance_check_row.append("")
            elif measured_current_tau >= calculated_tau_min and measured_current_tau <= calculated_tau_max:
                current_tolerance_check_row.append("PASS")
            else:
                current_tolerance_check_row.append("FAIL")

        with open(comparison_csv_path, mode="w", newline="") as csv_file:
            writer = csv.writer(csv_file)
            writer.writerow(headers)
            writer.writerow(calculated_nominal_row)
            writer.writerow(calculated_min_row)
            writer.writerow(calculated_max_row)
            writer.writerow(measured_voltage_row)
            writer.writerow(measured_current_row)
            writer.writerow(percent_difference_voltage_row)
            writer.writerow(percent_difference_current_row)
            writer.writerow(voltage_tolerance_check_row)
            writer.writerow(current_tolerance_check_row)

        if "rc_filter_model" in self.datastore:
            self.datastore["rc_filter_model"]["calculated_vs_measured_csv"] = comparison_csv_path

        print(f"{self.prefix} --> RC calculated-vs-measured summary CSV saved: {comparison_csv_path}")
    # ------------------------------------------------------------------------------------------------------
    def beep_sequence(self):
        #First beep is always longer for some reason
        self.r0.beep()
        time.sleep(1)

        self.r0.beep()
        time.sleep(0.5)
        self.r0.beep()
        time.sleep(0.125)
        self.r0.beep()
        time.sleep(0.125)
        self.r0.beep()
        time.sleep(0.5)
        self.r0.beep()
        time.sleep(0.75)

        self.k.beep()
        time.sleep(0.5)
        self.k.beep()

	

if __name__ == "__main__":
    if len(sys.argv) < 2:
        sys.exit(f"Error: You need to supply a config file for this test as the argument! You had {len(sys.argv)-1} arguments!")
    if (len(sys.argv) == 2):
        LDOmeasure(sys.argv[1])
    elif (len(sys.argv) == 3):
        LDOmeasure(sys.argv[1], sys.argv[2])
    else:
        sys.exit(f"Error: You need to supply a config file and optional test name for this program, 2 arguments max. You supplied {sys.argv}, which is {len(sys.argv)-1} arguments")
        
        
        
        
